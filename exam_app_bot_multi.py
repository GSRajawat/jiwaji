import streamlit as st
import pandas as pd
import os
import zipfile
import io
import fitz # PyMuPDF
import re
import tempfile
import ast
import requests
from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import json
from supabase import create_client, Client
import datetime
import numpy as np
import traceback
import csv


# --- Exam Center Configuration ---
# Define your exam centers with a display name and a unique identifier (ID)
# The ID will be used for file paths and Supabase filtering.
EXAM_CENTERS = {
    "Government Law College, Morena": "center_morena",
    "Government Engineering College, Gwalior": "center_gwalior",
    "MLIS College, Bhopal": "center_bhopal"
    # Add more centers as needed
}

# --- Initialize Supabase ---
try:
    SUPABASE_URL = st.secrets["supabase"]["url"]
    SUPABASE_KEY = st.secrets["supabase"]["key"]
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json"
    }
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
except KeyError:
    st.error("Supabase secrets not found. Please configure `supabase.url` and `supabase.key` in your secrets.toml file.")
    st.stop()


# --- Configuration ---
# These are now base filenames. The actual paths will be constructed dynamically.
CS_REPORTS_FILE = "cs_reports.csv"
EXAM_TEAM_MEMBERS_FILE = "exam_team_members.csv"
SHIFT_ASSIGNMENTS_FILE = "shift_assignments.csv"
ROOM_INVIGILATORS_FILE = "room_invigilator_assignments.csv"
SITTING_PLAN_FILE = "sitting_plan.csv"
TIMETABLE_FILE = "timetable.csv"
ASSIGNED_SEATS_FILE = "assigned_seats.csv"
ATTESTATION_DATA_FILE = "attestation_data_combined.csv"
COLLEGE_STATISTICS_FILE = "college_statistics_fancy.csv"


# --- Helper Function for Center-Specific File Paths ---
def get_center_filepath(filename):
    """
    Constructs a file path specific to the currently selected exam center.
    Files will be stored in 'data/{center_id}/{filename}'.
    """
    center_id = st.session_state.center_id
    base_dir = "data"
    center_dir = os.path.join(base_dir, center_id)
    os.makedirs(center_dir, exist_ok=True) # Ensure the directory exists
    return os.path.join(center_dir, filename)

def _format_paper_code(paper_code_str):
    """
    Standardizes paper codes by stripping whitespace, removing spaces,
    converting to uppercase, and removing non-alphanumeric characters
    except hyphens.
    """
    if pd.isna(paper_code_str):
        return ""
    code = str(paper_code_str).strip().replace(" ", "").upper()
    # Remove any non-alphanumeric characters except '-'
    code = re.sub(r'[^A-Z0-9-]', '', code)
    return code

# --- CORRECTED: upload_csv_to_supabase function ---
def upload_csv_to_supabase(table_name, csv_path, unique_cols=None):
    try:
        df = pd.read_csv(csv_path)
        df.columns = df.columns.str.strip()

        # --- IMPORTANT: Add center_id column before upload ---
        df['center_id'] = st.session_state.center_id

        # Column name mapping from CSV headers to database columns
        column_mappings = {
            # Common mappings
            'Roll Number': 'roll_number',
            'Paper Code': 'paper_code',
            'Paper Name': 'paper_name',
            'Room Number': 'room_number',
            'Seat Number': 'seat_number',
            'Date': 'date',
            'Shift': 'shift',
            'SN': 'sn',
            'Time': 'time',
            'Class': 'class',
            'Paper': 'paper',
            'Name': 'name',

            # Sitting plan specific
            'Roll Number 1': 'roll_number_1',
            'Roll Number 2': 'roll_number_2',
            'Roll Number 3': 'roll_number_3',
            'Roll Number 4': 'roll_number_4',
            'Roll Number 5': 'roll_number_5',
            'Roll Number 6': 'roll_number_6',
            'Roll Number 7': 'roll_number_7',
            'Roll Number 8': 'roll_number_8',
            'Roll Number 9': 'roll_number_9',
            'Roll Number 10': 'roll_number_10',
            'Mode': 'mode',
            'Type': 'type',
            'Seat Number 1': 'seat_number_1',
            'Seat Number 2': 'seat_number_2',
            'Seat Number 3': 'seat_number_3',
            'Seat Number 4': 'seat_number_4',
            'Seat Number 5': 'seat_number_5',
            'Seat Number 6': 'seat_number_6',
            'Seat Number 7': 'seat_number_7',
            'Seat Number 8': 'seat_number_8',
            'Seat Number 9': 'seat_number_9',
            'Seat Number 10': 'seat_number_10',

            # Attestation specific MAPPINGS ADDED/UPDATED HERE
            'Enrollment Number': 'enrollment_number',
            'Session': 'session',
            'Regular/Backlog': 'regular_backlog',
            'Father\'s Name': 'father_name',
            'Mother\'s Name': 'mother_name',
            'Gender': 'gender',
            'Exam Name': 'exam_name',
            'Exam Centre': 'exam_centre',
            'College Name': 'college_name',
            'Address': 'address',
            'Paper 1': 'paper_1',
            'Paper 2': 'paper_2',
            'Paper 3': 'paper_3',
            'Paper 4': 'paper_4',
            'Paper 5': 'paper_5',
            'Paper 6': 'paper_6',
            'Paper 7': 'paper_7',
            'Paper 8': 'paper_8',
            'Paper 9': 'paper_9',
            'Paper 10': 'paper_10',
            # Other specific mappings
            'report_key': 'report_key',
            'room_num': 'room_num',
            'absent_roll_numbers': 'absent_roll_numbers',
            'ufm_roll_numbers': 'ufm_roll_numbers',
            'invigilators': 'invigilators',
            'senior_center_superintendent': 'senior_center_superintendent',
            'center_superintendent': 'center_superintendent',
            'assistant_center_superintendent': 'assistant_center_superintendent',
            'permanent_invigilator': 'permanent_invigilator',
            'assistant_permanent_invigilator': 'assistant_permanent_invigilator',
            'class_3_worker': 'class_3_worker',
            'class_4_worker': 'class_4_worker',
            'center_id': 'center_id' # Add center_id mapping
        }
        
        # Rename columns according to mapping
        df.rename(columns=column_mappings, inplace=True)

        # Clean missing values and handle NaN
        df = df.replace(r'^\s*$', None, regex=True)
        df = df.replace([np.inf, -np.inf], None)

        # Convert all columns to handle NaN properly
        for col in df.columns:
            df[col] = df[col].apply(lambda x: None if pd.isna(x) or
                                                 (isinstance(x, float) and not np.isfinite(x)) or
                                                 str(x).strip() == '' else x)
        
        # --- CORRECTED CODE HERE ---
        # Handle date format conversion (DD-MM-YYYY to YYYY-MM-DD)
        if 'date' in df.columns:
            def convert_date_format(date_str):
                if pd.notna(date_str) and isinstance(date_str, str) and len(date_str.split('-')) == 3:
                    try:
                        # Assuming DD-MM-YYYY format from the CSV, convert to YYYY-MM-DD
                        return datetime.datetime.strptime(date_str, '%d-%m-%Y').strftime('%Y-%m-%d')
                    except ValueError:
                        # If the format is not as expected, return the original string
                        return date_str
                return date_str

            df['date'] = df['date'].apply(convert_date_format)

        # Handle JSON fields (arrays stored as strings)
        json_fields = ['absent_roll_numbers', 'ufm_roll_numbers', 'invigilators',
                     'senior_center_superintendent', 'center_superintendent',
                     'assistant_center_superintendent', 'permanent_invigilator',
                     'assistant_permanent_invigilator', 'class_3_worker', 'class_4_worker']
        
        for field in json_fields:
            if field in df.columns:
                def parse_json_field(x):
                    if pd.notna(x) and isinstance(x, str) and x.strip():
                        try:
                            # Safely evaluate if it's a list string
                            if x.strip().startswith('['):
                                return ast.literal_eval(x)
                            # If it's a single value, wrap it in a list
                            return [x.strip()]
                        except (ValueError, SyntaxError):
                            # Fallback if literal_eval fails
                            return [x.strip()]
                    return None
                df[field] = df[field].apply(parse_json_field)

        # Convert numeric fields properly
        numeric_fields = ['room_number', 'seat_number', 'room_num', 'sn'] + \
                         [f'seat_number_{i}' for i in range(1, 11)]

        for field in numeric_fields:
            if field in df.columns:
                df[field] = pd.to_numeric(df[field], errors='coerce')
                df[field] = df[field].astype('Int64')

        if df.empty:
            return False, f"⚠️ `{csv_path}` is empty."

        # Convert to records and ensure all NaN/None values are properly handled
        records = df.to_dict(orient='records')

        # Final cleanup of records
        cleaned_records = []
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                if isinstance(value, list):
                    cleaned_record[key] = value if value else None
                elif value is None or (not isinstance(value, list) and pd.isna(value)):
                    cleaned_record[key] = None
                elif isinstance(value, (np.int64, np.int32)):
                    cleaned_record[key] = int(value)
                elif isinstance(value, (np.float64, np.float32)):
                    cleaned_record[key] = float(value) if np.isfinite(value) else None
                else:
                    cleaned_record[key] = value
            cleaned_records.append(cleaned_record)

        # Upload in batches to handle large datasets
        batch_size = 100
        total_uploaded = 0

        for i in range(0, len(cleaned_records), batch_size):
            batch = cleaned_records[i:i + batch_size]
            supabase.table(table_name).insert(batch).execute()
            total_uploaded += len(batch)

        return True, f"✅ Uploaded {total_uploaded} rows to `{table_name}`."

    except Exception as e:
        return False, f"❌ Error uploading to `{table_name}`: {str(e)}"

# --- MODIFIED: download_supabase_to_csv (to handle API exceptions and center_id) ---
def download_supabase_to_csv(table_name, filename):
    all_data = []
    limit = 1000
    offset = 0

    try:
        while True:
            # Use supabase.from_ for explicit table selection with pagination and FILTER BY center_id
            response = supabase.from_(table_name).select("*").eq('center_id', st.session_state.center_id).limit(limit).offset(offset).execute()
            
            data = response.data
            
            # If no data is returned, we have reached the end of the table
            if not data:
                break
                
            all_data.extend(data)
            
            # If the number of records returned is less than the limit, we've reached the end
            if len(data) < limit:
                break
            
            offset += limit
    
    except Exception as e:
        # Catch any API exceptions (e.g., failed request, bad token)
        traceback.print_exc()
        return False, f"❌ Supabase API Error for '{table_name}': {e}"
    
    if not all_data:
        return True, f"⚠️ No data found in table `{table_name}` for center '{st.session_state.center_name}'. An empty file has been created."

    df = pd.DataFrame(all_data)
    
    # --- IMPORTANT: Remove 'id', 'created_at', and 'center_id' columns before saving to CSV ---
    columns_to_drop = ['id', 'created_at', 'center_id'] 
    df = df.drop(columns=[col for col in columns_to_drop if col in df.columns])
    
    # Reverse column name mapping (database columns back to CSV headers)
    reverse_column_mappings = {
        'roll_number': 'Roll Number', 'paper_code': 'Paper Code', 'paper_name': 'Paper Name',
        'room_number': 'Room Number', 'seat_number': 'Seat Number', 'date': 'Date',
        'shift': 'Shift', 'sn': 'SN', 'time': 'Time', 'class': 'Class', 'paper': 'Paper',
        'name': 'Name',
        'roll_number_1': 'Roll Number 1', 'roll_number_2': 'Roll Number 2',
        'roll_number_3': 'Roll Number 3', 'roll_number_4': 'Roll Number 4',
        'roll_number_5': 'Roll Number 5', 'roll_number_6': 'Roll Number 6',
        'roll_number_7': 'Roll Number 7', 'roll_number_8': 'Roll Number 8',
        'roll_number_9': 'Roll Number 9', 'roll_number_10': 'Roll Number 10',
        'mode': 'Mode', 'type': 'Type',
        'seat_number_1': 'Seat Number 1', 'seat_number_2': 'Seat Number 2',
        'seat_number_3': 'Seat Number 3', 'seat_number_4': 'Seat Number 4',
        'seat_number_5': 'Seat Number 5', 'seat_number_6': 'Seat Number 6',
        'seat_number_7': 'Seat Number 7', 'seat_number_8': 'Seat Number 8',
        'seat_number_9': 'Seat Number 9', 'seat_number_10': 'Seat Number 10',
        'enrollment_number': 'Enrollment Number',
        'session': 'Session',
        'regular_backlog': 'Regular/Backlog',
        'father_name': 'Father\'s Name',
        'mother_name': 'Mother\'s Name',
        'gender': 'Gender',
        'exam_name': 'Exam Name',
        'exam_centre': 'Exam Centre',
        'college_name': 'College Name',
        'address': 'Address',
        'paper_1': 'Paper 1', 'paper_2': 'Paper 2', 'paper_3': 'Paper 3',
        'paper_4': 'Paper 4', 'paper_5': 'Paper 5', 'paper_6': 'Paper 6',
        'paper_7': 'Paper 7', 'paper_8': 'Paper 8', 'paper_9': 'Paper 9',
        'paper_10': 'Paper 10',
        'report_key': 'report_key',
        'room_num': 'room_num',
        'absent_roll_numbers': 'absent_roll_numbers',
        'ufm_roll_numbers': 'ufm_roll_numbers',
        'invigilators': 'invigilators',
        'senior_center_superintendent': 'senior_center_superintendent',
        'center_superintendent': 'center_superintendent',
        'assistant_center_superintendent': 'assistant_center_superintendent',
        'permanent_invigilator': 'permanent_invigilator',
        'assistant_permanent_invigilator': 'assistant_permanent_invigilator',
        'class_3_worker': 'class_3_worker',
        'class_4_worker': 'class_4_worker'
    }
    actual_reverse_column_mappings = {k: v for k, v in reverse_column_mappings.items() if k in df.columns}
    df.rename(columns=actual_reverse_column_mappings, inplace=True)
    
    # Handle date format conversion (YYYY-MM-DD back to DD-MM-YYYY)
    if 'Date' in df.columns:
        def format_date_for_csv(d_str):
            if pd.isna(d_str) or not isinstance(d_str, str) or d_str.strip() == '':
                return ''
            try:
                dt_obj = datetime.datetime.strptime(d_str, '%Y-%m-%d')
                return dt_obj.strftime('%d-%m-%Y')
            except ValueError:
                return d_str
        df['Date'] = df['Date'].apply(format_date_for_csv)
    
    # Handle JSON fields back to string format
    json_fields_to_str = [
        'absent_roll_numbers', 'ufm_roll_numbers', 'invigilators',
        'senior_center_superintendent', 'center_superintendent',
        'assistant_center_superintendent', 'permanent_invigilator', 
        'assistant_permanent_invigilator', 'class_3_worker', 'class_4_worker'
    ]
    
    for field in json_fields_to_str:
        if field in df.columns:
            df[field] = df[field].apply(lambda x: str(x) if x is not None and x != [] else '')
    
    df = df.fillna('')
    df.to_csv(filename, index=False)
    
    return True, f"✅ Downloaded {len(df)} rows from `{table_name}` to `{filename}`."

    
# --- NEW FUNCTION: Download attestation_data_combined to parent folder (now center-specific) ---
def download_attestation_data_to_center_folder(): # Renamed for clarity
    # Use get_center_filepath to ensure correct path for attestation data
    output_filepath = get_center_filepath(ATTESTATION_DATA_FILE)
    
    st.info(f"Attempting to download '{ATTESTATION_DATA_FILE}' from Supabase to: {output_filepath}")
    success, message = download_supabase_to_csv("attestation_data_combined", output_filepath)
    
    if success:
        st.success(f"Successfully downloaded '{ATTESTATION_DATA_FILE}' to the current center's data folder.")
        # st.rerun() # Reloading here might interrupt user flow. Only reruns if absolutely necessary.
    else:
        st.error(f"Failed to download '{ATTESTATION_DATA_FILE}': {message}")
    
    return success, message

# --- MODIFIED: load_data (remove attestation download from here, use center-specific paths) ---
def load_data():
    sitting_plan_df = pd.DataFrame()
    timetable_df = pd.DataFrame()
    assigned_seats_df = pd.DataFrame(columns=["Roll Number", "Paper Code", "Paper Name", "Room Number", "Seat Number", "Date", "Shift"])
    attestation_df = pd.DataFrame()

    # --- Initial Downloads from Supabase if local files are missing/empty ---
    
    # Download Timetable
    timetable_filepath = get_center_filepath(TIMETABLE_FILE)
    if not os.path.exists(timetable_filepath) or os.stat(timetable_filepath).st_size == 0:
        st.info(f"Attempting to download {TIMETABLE_FILE} from Supabase for '{st.session_state.center_name}'...")
        success, message = download_supabase_to_csv("timetable", timetable_filepath)
        if not success:
            st.warning(f"Failed to download {TIMETABLE_FILE} from Supabase: {message}. Will proceed with an empty DataFrame.")
    
    # Download Sitting Plan
    sitting_plan_filepath = get_center_filepath(SITTING_PLAN_FILE)
    if not os.path.exists(sitting_plan_filepath) or os.stat(sitting_plan_filepath).st_size == 0:
        st.info(f"Attempting to download {SITTING_PLAN_FILE} from Supabase for '{st.session_state.center_name}'...")
        success, message = download_supabase_to_csv("sitting_plan", sitting_plan_filepath)
        if not success:
            st.warning(f"Failed to download {SITTING_PLAN_FILE} from Supabase: {message}. Will proceed with an empty DataFrame.")

    # Download Assigned Seats
    assigned_seats_filepath = get_center_filepath(ASSIGNED_SEATS_FILE)
    if not os.path.exists(assigned_seats_filepath) or os.stat(assigned_seats_filepath).st_size == 0:
        st.info(f"Attempting to download {ASSIGNED_SEATS_FILE} from Supabase for '{st.session_state.center_name}'...")
        success, message = download_supabase_to_csv("assigned_seats", assigned_seats_filepath)
        if not success:
            st.warning(f"Failed to download {ASSIGNED_SEATS_FILE} from Supabase: {message}. Will proceed with an empty DataFrame.")

    # NOTE: Attestation data download logic is moved to its own function
    # You will need to call `download_attestation_data_to_center_folder()` explicitly
    # if you want this file to be downloaded by a user action.
            
    # Load all DataFrames from local CSVs (now potentially updated from Supabase)
    if os.path.exists(sitting_plan_filepath):
        try:
            sitting_plan_df = pd.read_csv(sitting_plan_filepath, dtype={
                f"Roll Number {i}": str for i in range(1, 11)
            })
            sitting_plan_df.columns = sitting_plan_df.columns.str.strip().str.replace('\ufeff', '').str.replace('\xa0', ' ')
            if 'Paper Code' in sitting_plan_df.columns:
                sitting_plan_df['Paper Code'] = sitting_plan_df['Paper Code'].apply(_format_paper_code)
        except Exception as e:
            st.error(f"Error loading {SITTING_PLAN_FILE} from local file: {e}")
            sitting_plan_df = pd.DataFrame()

    if os.path.exists(timetable_filepath):
        try:
            timetable_df = pd.read_csv(timetable_filepath)
            timetable_df.columns = timetable_df.columns.str.strip().str.replace('\ufeff', '').str.replace('\xa0', ' ')
            if 'Paper Code' in timetable_df.columns:
                timetable_df['Paper Code'] = timetable_df['Paper Code'].apply(_format_paper_code)
        except Exception as e:
            st.error(f"Error loading {TIMETABLE_FILE} from local file: {e}")
            timetable_df = pd.DataFrame()
    
    if os.path.exists(assigned_seats_filepath):
        try:
            temp_assigned_df = pd.read_csv(assigned_seats_filepath, dtype=str)
            temp_assigned_df.columns = temp_assigned_df.columns.str.strip().str.replace('\ufeff', '').str.replace('\xa0', ' ')
            rename_map = {}
            # More robust column renaming for common typos
            if 'Roll Numb' in temp_assigned_df.columns: rename_map['Roll Numb'] = 'Roll Number'
            if 'Paper Cod' in temp_assigned_df.columns: rename_map['Paper Cod'] = 'Paper Code'
            if 'Paper Nan' in temp_assigned_df.columns: rename_map['Paper Nan'] = 'Paper Name'
            if 'Room Nur' in temp_assigned_df.columns: rename_map['Room Nur'] = 'Room Number'
            if 'Seat Numi' in temp_assigned_df.columns: rename_map['Seat Numi'] = 'Seat Number'
            
            # Standardize 'Date' and 'Shift' to lowercase if they exist
            if 'Date' in temp_assigned_df.columns: rename_map['Date'] = 'date'
            if 'Shift' in temp_assigned_df.columns: rename_map['Shift'] = 'shift'

            if rename_map:
                temp_assigned_df.rename(columns=rename_map, inplace=True)
            
            # Ensure 'date' and 'shift' are lowercase for consistency later
            temp_assigned_df.columns = temp_assigned_df.columns.str.lower()


            required_assigned_cols=["roll number","paper code","paper name","room number","seat number","date","shift"]
            missing_cols=[col for col in required_assigned_cols if col not in temp_assigned_df.columns]
            if missing_cols:
                st.error(f"Critical Error: Missing essential columns in {ASSIGNED_SEATS_FILE}: {missing_cols}. Please verify the file content.")
                assigned_seats_df=pd.DataFrame(columns=required_assigned_cols)
            else:
                assigned_seats_df=temp_assigned_df[required_assigned_cols].copy()
                assigned_seats_df['paper code']=assigned_seats_df['paper code'].apply(_format_paper_code) # Ensure paper code is formatted
                for col in ["roll number","room number","seat number","date","shift"]:
                    if col in assigned_seats_df.columns:
                        assigned_seats_df[col]=assigned_seats_df[col].astype(str) # Ensure string type for merging/comparison
        except Exception as e:
            st.error(f"Error loading {ASSIGNED_SEATS_FILE} from local file: {e}. Ensure it's a valid CSV file.")
            assigned_seats_df=pd.DataFrame(columns=["Roll Number","Paper Code","Paper Name","Room Number","Seat Number","Date","Shift"])
    
    # Load attestation data from the center-specific folder
    attestation_filepath = get_center_filepath(ATTESTATION_DATA_FILE)
    if os.path.exists(attestation_filepath):
        try:
            attestation_df=pd.read_csv(attestation_filepath,dtype=str)
            attestation_df.columns=attestation_df.columns.str.strip().str.replace('\ufeff','').str.replace('\xa0',' ')
            for i in range(1,11):
                col_name=f'Paper {i}'
                if col_name in attestation_df.columns:
                    attestation_df[col_name]=attestation_df[col_name].fillna('').astype(str)
            st.info(f"Loaded '{ATTESTATION_DATA_FILE}' from current center's data folder.")
        except Exception as e:
            st.error(f"Error loading {ATTESTATION_DATA_FILE} from current center's data folder: {e}. Ensure it's a valid CSV file.")
            attestation_df=pd.DataFrame()
    else:
        st.warning(f"'{ATTESTATION_DATA_FILE}' not found in the current center's data folder. Some features may be limited. Consider downloading it from Supabase.")


    st.session_state['sitting_plan']=sitting_plan_df
    st.session_state['timetable']=timetable_df
    st.session_state['assigned_seats_df']=assigned_seats_df
    st.session_state['attestation_df']=attestation_df
    return sitting_plan_df,timetable_df,assigned_seats_df,attestation_df

def load_shift_assignments():
    file_path = get_center_filepath(SHIFT_ASSIGNMENTS_FILE)
    if os.path.exists(file_path):
        try:
            df = pd.read_csv(file_path, engine='python')
            df.columns = df.columns.str.lower() # Standardize to lowercase
            def safe_literal_eval(val):
                if isinstance(val, str) and val.strip():
                    clean_val = val.strip().strip('"') # Clean extra quotes
                    try:
                        return ast.literal_eval(clean_val)
                    except (ValueError, SyntaxError):
                        return [clean_val] # Return as single-item list if parse fails
                return []
            for role in ["senior_center_superintendent","center_superintendent","assistant_center_superintendent","permanent_invigilator","assistant_permanent_invigilator","class_3_worker","class_4_worker"]:
                if role in df.columns:
                    df[role] = df[role].apply(safe_literal_eval)
            return df
        except Exception as e:
            st.error(f"Error loading shift assignments: {e}. Returning empty DataFrame.")
            return pd.DataFrame(columns=['date','shift','senior_center_superintendent','center_superintendent',"assistant_center_superintendent","permanent_invigilator","assistant_permanent_invigilator","class_3_worker","class_4_worker"])
    return pd.DataFrame(columns=['date','shift','senior_center_superintendent','center_superintendent',"assistant_center_superintendent","permanent_invigilator","assistant_permanent_invigilator","class_3_worker","class_4_worker"])

def save_shift_assignment(date, shift, assignments):
    assignments_df = load_shift_assignments()
    assignment_key = f"{date}_{shift}"
    data_for_df = {
        'date': date,
        'shift': shift,
        'senior_center_superintendent': str(assignments.get('senior_center_superintendent', [])),
        'center_superintendent': str(assignments.get('center_superintendent', [])),
        'assistant_center_superintendent': str(assignments.get('assistant_center_superintendent', [])),
        'permanent_invigilator': str(assignments.get('permanent_invigilator', [])),
        'assistant_permanent_invigilator': str(assignments.get('assistant_permanent_invigilator', [])),
        'class_3_worker': str(assignments.get('class_3_worker', [])),
        'class_4_worker': str(assignments.get('class_4_worker', []))
    }
    new_row_df = pd.DataFrame([data_for_df])

    if assignment_key in (assignments_df['date'] + '_' + assignments_df['shift']).values:
        idx_to_update = assignments_df[(assignments_df['date'] == date) & (assignments_df['shift'] == shift)].index[0]
        for col, val in data_for_df.items():
            assignments_df.loc[idx_to_update, col] = val
    else:
        assignments_df = pd.concat([assignments_df, new_row_df], ignore_index=True)
    
    try:
        assignments_df.to_csv(get_center_filepath(SHIFT_ASSIGNMENTS_FILE), index=False)
        return True, "Shift assignments saved successfully!"
    except Exception as e:
        return False, f"Error saving shift assignments: {e}"

def save_uploaded_file(uploaded_file_content, filename):
    try:
        file_path = get_center_filepath(filename) # Use center-specific path
        if isinstance(uploaded_file_content, pd.DataFrame):
            # For DataFrames, convert to CSV bytes
            csv_bytes = uploaded_file_content.to_csv(index=False).encode('utf-8')
        else:
            # For uploaded file objects, read bytes
            if hasattr(uploaded_file_content, 'getbuffer'): # Streamlit UploadedFile
                csv_bytes = uploaded_file_content.getbuffer()
            else: # Other file-like objects
                csv_bytes = uploaded_file_content.read()

        with open(file_path, "wb") as f:
            f.write(csv_bytes)
        return True, f"File {filename} saved successfully!"
    except Exception as e:
        return False, f"Error saving file {filename}: {e}"


def admin_login():
    user = st.text_input("Username", type="default", key="admin_user")
    pwd = st.text_input("Password", type="password", key="admin_pass")
    return user == "admin" and pwd == "admin123"

def cs_login():
    user = st.text_input("CS Username", type="default", key="cs_user")
    pwd = st.text_input("CS Password", type="password", key="cs_pass")
    return user == "cs_admin" and pwd == "cs_pass123"

def load_cs_reports_csv():
    file_path = get_center_filepath(CS_REPORTS_FILE)
    if os.path.exists(file_path):
        try:
            df = pd.read_csv(file_path)
            # Standardize column names to lowercase and replace spaces
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            if 'class' not in df.columns:
                df['class'] = "" # Ensure 'class' column exists
            for col in ['absent_roll_numbers', 'ufm_roll_numbers']:
                if col in df.columns:
                    df[col] = df[col].astype(str).apply(lambda x: ast.literal_eval(x) if x.strip() and x.strip().lower() != 'nan' else [])
            return df
        except Exception as e:
            st.error(f"Error loading CS reports from CSV: {e}")
            return pd.DataFrame(columns=['report_key', 'date', 'shift', 'room_num', 'paper_code', 'paper_name', 'class', 'absent_roll_numbers', 'ufm_roll_numbers'])
    else:
        # Return a DataFrame with all expected columns if the file doesn't exist
        return pd.DataFrame(columns=['report_key', 'date', 'shift', 'room_num', 'paper_code', 'paper_name', 'class', 'absent_roll_numbers', 'ufm_roll_numbers'])

def save_cs_report_csv(report_key, data):
    reports_df = load_cs_reports_csv()
    data_for_df = data.copy()
    # Ensure list fields are stored as strings for CSV compatibility
    data_for_df['absent_roll_numbers'] = str(data_for_df.get('absent_roll_numbers', []))
    data_for_df['ufm_roll_numbers'] = str(data_for_df.get('ufm_roll_numbers', []))

    new_row_df = pd.DataFrame([data_for_df])

    if report_key in reports_df['report_key'].values:
        # Update existing row
        idx_to_update = reports_df[reports_df['report_key'] == report_key].index[0]
        for col, val in data_for_df.items():
            reports_df.loc[idx_to_update, col] = val
    else:
        # Add new row
        reports_df = pd.concat([reports_df, new_row_df], ignore_index=True)
    
    try:
        reports_df.to_csv(get_center_filepath(CS_REPORTS_FILE), index=False)
        return True, "Report saved to CSV successfully!"
    except Exception as e:
        return False, f"Error saving report to CSV: {e}"

def load_single_cs_report_csv(report_key):
    reports_df = load_cs_reports_csv()
    filtered_df = reports_df[reports_df['report_key'] == report_key]
    if not filtered_df.empty:
        return True, filtered_df.iloc[0].to_dict()
    else:
        return False, {}

def load_exam_team_members():
    file_path = get_center_filepath(EXAM_TEAM_MEMBERS_FILE)
    if os.path.exists(file_path):
        try:
            df = pd.read_csv(file_path)
            return df['Name'].tolist()
        except Exception as e:
            st.error(f"Error loading exam team members: {e}")
            return []
    return []

def save_exam_team_members(members):
    df = pd.DataFrame({'Name': sorted(list(set(members)))})
    try:
        df.to_csv(get_center_filepath(EXAM_TEAM_MEMBERS_FILE), index=False)
        return True, "Exam team members saved successfully!"
    except Exception as e:
        return False, f"Error saving exam team members: {e}"

def _get_session_students_raw_data(date_str, shift, assigned_seats_df, timetable_df):
    all_students_data = []
    # Ensure date and shift columns are strings for comparison
    current_day_exams_tt = timetable_df[(timetable_df["Date"].astype(str).str.strip() == date_str) & 
                                        (timetable_df["Shift"].astype(str).str.strip().str.lower() == shift.lower())].copy()
    if current_day_exams_tt.empty:
        return all_students_data

    for _, tt_row in current_day_exams_tt.iterrows():
        tt_class = str(tt_row["Class"]).strip()
        tt_paper_code = str(tt_row["Paper Code"]).strip()
        tt_paper_name = str(tt_row["Paper Name"]).strip()

        # Ensure 'date' and 'shift' in assigned_seats_df are lowercase before filtering
        assigned_seats_df.columns = assigned_seats_df.columns.str.lower() 

        current_exam_assigned_students = assigned_seats_df[
            (assigned_seats_df["date"].astype(str).str.strip() == date_str) &
            (assigned_seats_df["shift"].astype(str).str.strip().str.lower() == shift.lower()) &
            (assigned_seats_df["paper code"].astype(str).str.strip() == tt_paper_code) &
            (assigned_seats_df["paper name"].astype(str).str.strip() == tt_paper_name)
        ]

        for _, assigned_row in current_exam_assigned_students.iterrows():
            roll_num = str(assigned_row["roll number"]).strip()
            room_num = str(assigned_row["room number"]).strip()
            seat_num_raw = str(assigned_row["seat number"]).strip()

            seat_num_display = ""
            seat_num_sort_key = None
            try:
                if re.match(r'^\d+[A-Z]$', seat_num_raw):
                    num_part = int(re.match(r'^(\d+)', seat_num_raw).group(1))
                    char_part = re.search(r'([A-Z])$', seat_num_raw).group(1)
                    seat_num_sort_key = (ord(char_part), num_part)
                    seat_num_display = seat_num_raw
                elif seat_num_raw.isdigit():
                    seat_num_sort_key = (float('inf'), int(seat_num_raw))
                    seat_num_display = str(int(float(seat_num_raw)))
                else:
                    seat_num_sort_key = (float('inf'), float('inf'))
                    seat_num_display = seat_num_raw if seat_num_raw else "N/A"
            except ValueError:
                seat_num_sort_key = (float('inf'), float('inf'))
                seat_num_display = seat_num_raw if seat_num_raw else "N/A"

            all_students_data.append({
                "roll_num": roll_num,
                "room_num": room_num,
                "seat_num_display": seat_num_display,
                "seat_num_sort_key": seat_num_sort_key,
                "paper_name": tt_paper_name,
                "paper_code": tt_paper_code,
                "class_name": tt_class,
                "date": date_str,
                "shift": shift
            })
    return all_students_data

def get_all_students_for_date_shift_formatted(date_str, shift, assigned_seats_df, timetable):
    all_students_data = _get_session_students_raw_data(date_str, shift, assigned_seats_df, timetable)
    if not all_students_data:
        return None, "No students found for the selected date and shift.", None

    all_students_data.sort(key=lambda x: (x['room_num'], x['seat_num_sort_key']))

    current_day_exams_tt = timetable[(timetable["Date"].astype(str).str.strip() == date_str) & 
                                     (timetable["Shift"].astype(str).str.strip().str.lower() == shift.lower())]
    
    exam_time = current_day_exams_tt.iloc[0]["Time"].strip() if "Time" in current_day_exams_tt.columns and not current_day_exams_tt.empty else "TBD"

    unique_classes = current_day_exams_tt['Class'].dropna().astype(str).str.strip().unique()
    class_summary_header = ""
    if len(unique_classes) == 1:
        class_summary_header = f"{unique_classes[0]} Examination {datetime.datetime.now().year}"
    elif len(unique_classes) > 1:
        class_summary_header = f"Various Classes Examination {datetime.datetime.now().year}"
    else:
        class_summary_header = f"Examination {datetime.datetime.now().year}" # Fallback if no classes found

    output_string_parts = []
    output_string_parts.append("जीवाजी विश्वविद्यालय ग्वालियर")
    output_string_parts.append(f"परीक्षा केंद्र :- {st.session_state.center_name} कोड :- {st.session_state.center_id.upper()}")
    output_string_parts.append(class_summary_header)
    output_string_parts.append(f"दिनांक :-{date_str}")
    output_string_parts.append(f"पाली :-{shift}")
    output_string_parts.append(f"समय :-{exam_time}")

    students_by_room = {}
    for student in all_students_data:
        room = student['room_num']
        if room not in students_by_room:
            students_by_room[room] = []
        students_by_room[room].append(student)

    for room_num in sorted(students_by_room.keys()):
        output_string_parts.append(f" कक्ष :-{room_num}")
        current_room_students = students_by_room[room_num]
        num_cols = 10  # Number of students per line in text output
        for i in range(0, len(current_room_students), num_cols):
            block_students = current_room_students[i:i + num_cols]
            single_line_students = []
            for student in block_students:
                single_line_students.append(f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']})-{student['paper_name']}")
            output_string_parts.append(" ".join(single_line_students)) # Join with space for readability

    final_text_output = "\n".join(output_string_parts)

    excel_output_data = []
    excel_output_data.append(["जीवाजी विश्वविद्यालय ग्वालियर"])
    excel_output_data.append([f"परीक्षा केंद्र :- {st.session_state.center_name} कोड :- {st.session_state.center_id.upper()}"])
    excel_output_data.append([class_summary_header])
    excel_output_data.append([])
    excel_output_data.append(["दिनांक :-", date_str])
    excel_output_data.append(["पाली :-", shift])
    excel_output_data.append(["समय :-", exam_time])
    excel_output_data.append([])

    for room_num in sorted(students_by_room.keys()):
        excel_output_data.append([f" कक्ष :-{room_num}"])
        current_room_students = students_by_room[room_num]
        num_cols = 10
        for i in range(0, len(current_room_students), num_cols):
            block_students = current_room_students[i:i + num_cols]
            excel_row_for_students = [""] * num_cols # Initialize with empty strings
            for k, student in enumerate(block_students):
                excel_row_for_students[k] = (f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']})-{student['paper_name']}")
            excel_output_data.append(excel_row_for_students)
            excel_output_data.append([""] * num_cols) # Empty row for spacing in Excel

    return final_text_output, None, excel_output_data


def load_room_invigilator_assignments():
    file_path = get_center_filepath(ROOM_INVIGILATORS_FILE)
    if os.path.exists(file_path):
        try:
            df = pd.read_csv(file_path)
            df.columns = df.columns.str.lower() # Standardize to lowercase
            if 'invigilators' in df.columns:
                df['invigilators'] = df['invigilators'].astype(str).apply(lambda x: ast.literal_eval(x) if x.strip() and x.strip().lower() != 'nan' else [])
            return df
        except Exception as e:
            st.error(f"Error loading room invigilator assignments: {e}")
            return pd.DataFrame(columns=['date', 'shift', 'room_num', 'invigilators'])
    return pd.DataFrame(columns=['date', 'shift', 'room_num', 'invigilators'])

def save_room_invigilator_assignment(date, shift, room_num, invigilators):
    inv_df = load_room_invigilator_assignments()
    assignment_key = f"{date}_{shift}_{room_num}"
    data_for_df = {
        'date': date,
        'shift': shift,
        'room_num': room_num,
        'invigilators': str(invigilators) # Store list as string
    }
    new_row_df = pd.DataFrame([data_for_df])

    # Check if a record for this date, shift, and room already exists
    if assignment_key in (inv_df['date'] + '_' + inv_df['shift'] + '_' + inv_df['room_num'].astype(str)).values:
        idx_to_update = inv_df[(inv_df['date'] == date) & 
                               (inv_df['shift'] == shift) & 
                               (inv_df['room_num'].astype(str) == str(room_num))].index[0]
        for col, val in data_for_df.items():
            inv_df.loc[idx_to_update, col] = val
    else:
        inv_df = pd.concat([inv_df, new_row_df], ignore_index=True)
    
    try:
        inv_df.to_csv(get_center_filepath(ROOM_INVIGILATORS_FILE), index=False)
        return True, "Room invigilator assignments saved successfully!"
    except Exception as e:
        return False, f"Error saving room invigilator assignments: {e}"

def get_all_exams(roll_number, sitting_plan, timetable):
    student_exams = []
    roll_number_str = str(roll_number).strip()

    for _, sp_row in sitting_plan.iterrows():
        for i in range(1, 11):
            r_col = f"Roll Number {i}"
            if r_col in sp_row and pd.notna(sp_row[r_col]) and str(sp_row[r_col]).strip() == roll_number_str:
                paper = str(sp_row["Paper"]).strip()
                paper_code = str(sp_row["Paper Code"]).strip()
                paper_name = str(sp_row["Paper Name"]).strip()
                _class = str(sp_row["Class"]).strip()

                matches_in_timetable = timetable[
                    (timetable["Paper"].astype(str).str.strip() == paper) &
                    (timetable["Paper Code"].astype(str).str.strip() == paper_code) &
                    (timetable["Paper Name"].astype(str).str.strip() == paper_name) &
                    (timetable["Class"].astype(str).str.strip().str.lower() == _class.lower())
                ]
                for _, tt_row in matches_in_timetable.iterrows():
                    student_exams.append({
                        "Date": tt_row["Date"],
                        "Shift": tt_row["Shift"],
                        "Class": _class,
                        "Paper": paper,
                        "Paper Code": paper_code,
                        "Paper Name": paper_name
                    })
                break # Break from inner loop once roll number is found in a sitting plan row
    return student_exams

def get_sitting_details(roll_number, date, sitting_plan, timetable):
    found_sittings = []
    roll_number_str = str(roll_number).strip()
    date_str = str(date).strip() # Ensure date is string for comparison

    for _, sp_row in sitting_plan.iterrows():
        for i in range(1, 11): # Check all 10 roll number columns
            r_col = f"Roll Number {i}"
            s_col = f"Seat Number {i}"
            if r_col in sp_row and pd.notna(sp_row[r_col]) and str(sp_row[r_col]).strip() == roll_number_str:
                paper = str(sp_row["Paper"]).strip()
                paper_code = str(sp_row["Paper Code"]).strip()
                paper_name = str(sp_row["Paper Name"]).strip()
                _class = str(sp_row["Class"]).strip()

                matches_in_timetable = timetable[
                    (timetable["Class"].astype(str).str.strip().str.lower() == _class.lower()) &
                    (timetable["Paper"].astype(str).str.strip() == paper) &
                    (timetable["Paper Code"].astype(str).str.strip() == paper_code) &
                    (timetable["Paper Name"].astype(str).str.strip() == paper_name) &
                    (timetable["Date"].astype(str).str.strip() == date_str) # Match date
                ]
                if not matches_in_timetable.empty:
                    for _, tt_row in matches_in_timetable.iterrows():
                        seat_num_display = ""
                        seat_num_sort_key = float('inf') # Default sort key

                        if s_col in sp_row.index:
                            seat_num_raw = str(sp_row[s_col]).strip()
                            try:
                                # Handle alphanumeric seats like 1A, 1B
                                if re.match(r'^\d+[A-Z]$', seat_num_raw):
                                    num_part = int(re.match(r'^(\d+)', seat_num_raw).group(1))
                                    char_part = re.search(r'([A-Z])$', seat_num_raw).group(1)
                                    seat_num_sort_key = (ord(char_part), num_part)
                                    seat_num_display = seat_num_raw
                                elif seat_num_raw.isdigit(): # Numeric seats
                                    seat_num_sort_key = (float('inf'), int(float(seat_num_raw))) # Use inf to put numeric after alphanumeric
                                    seat_num_display = str(int(float(seat_num_raw)))
                                else:
                                    seat_num_display = seat_num_raw if seat_num_raw else "N/A"
                            except ValueError:
                                seat_num_display = seat_num_raw if seat_num_raw else "N/A" # Fallback
                        else:
                            seat_num_display = "N/A" # If seat column doesn't exist

                        found_sittings.append({
                            "Room Number": sp_row["Room Number"],
                            "Seat Number": seat_num_display,
                            "Class": _class,
                            "Paper": paper,
                            "Paper Code": paper_code,
                            "Paper Name": paper_name,
                            "Date": tt_row["Date"],
                            "Shift": tt_row["Shift"],
                            "Mode": sp_row.get("Mode", ""),
                            "Type": sp_row.get("Type", "")
                        })
    return found_sittings

def get_all_students_roll_number_wise_formatted(date_str, shift, assigned_seats_df, timetable):
    all_students_data = _get_session_students_raw_data(date_str, shift, assigned_seats_df, timetable)
    if not all_students_data:
        return None, "No students found for the selected date and shift.", None

    all_students_data.sort(key=lambda x: x['roll_num']) # Sort by roll number

    current_day_exams_tt = timetable[(timetable["Date"].astype(str).str.strip() == date_str) & 
                                     (timetable["Shift"].astype(str).str.strip().str.lower() == shift.lower())]
    
    exam_time = current_day_exams_tt.iloc[0]["Time"].strip() if "Time" in current_day_exams_tt.columns and not current_day_exams_tt.empty else "TBD"

    unique_classes = current_day_exams_tt['Class'].dropna().astype(str).str.strip().unique()
    class_summary_header = ""
    if len(unique_classes) == 1:
        class_summary_header = f"{unique_classes[0]} Examination {datetime.datetime.now().year}"
    elif len(unique_classes) > 1:
        class_summary_header = f"Various Classes Examination {datetime.datetime.now().year}"
    else:
        class_summary_header = f"Examination {datetime.datetime.now().year}" # Fallback

    output_string_parts = []
    output_string_parts.append("जीवाजी विश्वविद्यालय ग्वालियर")
    output_string_parts.append(f"परीक्षा केंद्र :- {st.session_state.center_name} कोड :- {st.session_state.center_id.upper()}")
    output_string_parts.append(class_summary_header)
    output_string_parts.append(f"दिनांक :-{date_str}")
    output_string_parts.append(f"पाली :-{shift}")
    output_string_parts.append(f"समय :-{exam_time}")
    output_string_parts.append("") # Spacer line

    num_cols = 10 # Students per line in text output
    for i in range(0, len(all_students_data), num_cols):
        block_students = all_students_data[i:i + num_cols]
        single_line_students = []
        for student in block_students:
            single_line_students.append(f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']}){student['paper_name']}")
        output_string_parts.append(" ".join(single_line_students)) # Join with space for readability

    final_text_output = "\n".join(output_string_parts)

    excel_output_data = []
    excel_output_data.append(["जीवाजी विश्वविद्यालय ग्वालियर"])
    excel_output_data.append([f"परीक्षा केंद्र :- {st.session_state.center_name} कोड :- {st.session_state.center_id.upper()}"])
    excel_output_data.append([class_summary_header])
    excel_output_data.append([]) # Spacer row
    excel_output_data.append(["दिनांक :-", date_str])
    excel_output_data.append(["पाली :-", shift])
    excel_output_data.append(["समय :-", exam_time])
    excel_output_data.append([]) # Spacer row

    for i in range(0, len(all_students_data), num_cols):
        block_students = all_students_data[i:i + num_cols]
        excel_row_for_students = [""] * num_cols
        for k, student in enumerate(block_students):
            excel_row_for_students[k] = (f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']}){student['paper_name']}")
        excel_output_data.append(excel_row_for_students)
        excel_output_data.append([""] * num_cols) # Empty row for spacing in Excel

    return final_text_output, None, excel_output_data

def extract_metadata_from_pdf_text(text):
    # This function is used by process_sitting_plan_pdfs
    class_match = re.search(r'([A-Z]+)\s*/?\s*(\d+(SEM|YEAR))', text)
    class_val = f"{class_match.group(1)} {class_match.group(2)}" if class_match else "UNSPECIFIED_CLASS"

    pattern_match = re.search(r'([A-Z]+)\s*/\s*(\d+(?:SEM|YEAR))\s*/\s*([A-Z]+)\s*/\s*([A-Z]+)\s*/\s*MAR-2025', text)
    if pattern_match:
        mode_type = pattern_match.group(3)
        type_type = pattern_match.group(4)
    else:
        # Fallback if pattern_match doesn't work
        mode_type = "UNSPECIFIED_MODE"
        for keyword_mode in ["PRIVATE", "REGULAR"]:
            if keyword_mode in text.upper():
                mode_type = keyword_mode
                break
        type_type = "UNSPECIFIED_TYPE"
        for keyword_type in ["ATKT", "SUPP", "EXR", "REGULAR", "PRIVATE"]:
            if keyword_type in text.upper():
                type_type = keyword_type
                break

    paper_code = re.search(r'Paper Code[:\s]*([A-Z0-9]+)', text, re.IGNORECASE)
    paper_code_val = _format_paper_code(paper_code.group(1)) if paper_code else "UNSPECIFIED_PAPER_CODE"

    paper_name = re.search(r'Paper Name[:\s]*(.+?)(?:\n|$)', text)
    paper_name_val = paper_name.group(1).strip() if paper_name else "UNSPECIFIED_PAPER_NAME"

    return {"class": class_val, "mode": mode_type, "type": type_type, 
            "room_number": "", "seat_numbers": [""] * 10, 
            "paper_code": paper_code_val, "paper_name": paper_name_val}


def process_sitting_plan_pdfs(zip_file_buffer, output_sitting_plan_path, output_timetable_path):
    all_rows = []
    sitting_plan_columns = [f"Roll Number {i+1}" for i in range(10)]
    sitting_plan_columns += ["Class", "Mode", "Type", "Room Number"]
    sitting_plan_columns += [f"Seat Number {i+1}" for i in range(10)]
    sitting_plan_columns += ["Paper", "Paper Code", "Paper Name"] # Add Paper Name

    def extract_roll_numbers(text):
        return sorted(list(set(re.findall(r'\b\d{9}\b', text)))) # Extracts 9-digit roll numbers

    def format_sitting_plan_rows(rolls, paper_folder_name, meta):
        rows = []
        for i in range(0, len(rolls), 10): # Group rolls into blocks of 10
            row = rolls[i:i+10]
            while len(row) < 10: # Pad with empty strings if less than 10
                row.append("")
            
            # Add metadata to the row
            row.extend([meta["class"], meta["mode"], meta["type"], meta["room_number"]])
            row.extend(meta["seat_numbers"]) # Add placeholder seat numbers for now
            row.append(paper_folder_name) # Assuming folder name can be paper name
            row.append(meta["paper_code"])
            row.append(meta["paper_name"]) # Add paper name
            rows.append(row)
        return rows

    unique_exams_for_timetable = [] # To collect unique exam details for timetable
    
    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_file_buffer, 'r') as zip_ref:
            zip_ref.extractall(tmpdir)

        base_dir = tmpdir
        # Check if there's a single subdirectory and use that as base_dir
        extracted_contents = os.listdir(tmpdir)
        if 'pdf_folder' in extracted_contents and os.path.isdir(os.path.join(tmpdir, 'pdf_folder')):
            base_dir = os.path.join(tmpdir, 'pdf_folder')
        elif len(extracted_contents) == 1 and os.path.isdir(os.path.join(tmpdir, extracted_contents[0])):
            base_dir = os.path.join(tmpdir, extracted_contents[0])

        processed_files_count = 0
        for folder_name in os.listdir(base_dir):
            folder_path = os.path.join(base_dir, folder_name)
            if os.path.isdir(folder_path):
                for file in os.listdir(folder_path):
                    if file.lower().endswith(".pdf"):
                        pdf_path = os.path.join(folder_path, file)
                        try:
                            doc = fitz.open(pdf_path)
                            full_text = "\n".join(page.get_text() for page in doc)
                            doc.close()

                            current_meta = extract_metadata_from_pdf_text(full_text)
                            # Use folder_name as fallback for paper code/name if not found in PDF
                            if current_meta['paper_code'] == "UNSPECIFIED_PAPER_CODE":
                                current_meta['paper_code'] = folder_name
                            if current_meta['paper_name'] == "UNSPECIFIED_PAPER_NAME":
                                current_meta['paper_name'] = folder_name

                            rolls = extract_roll_numbers(full_text)
                            rows = format_sitting_plan_rows(rolls, paper_folder_name=folder_name, meta=current_meta)
                            all_rows.extend(rows)
                            processed_files_count += 1
                            st.info(f"✔ Processed: {file} ({len(rolls)} unique roll numbers)")

                            # Collect unique exam details for timetable generation
                            unique_exams_for_timetable.append({
                                'Class': current_meta['class'],
                                'Paper': folder_name, # Use folder name as paper identifier
                                'Paper Code': current_meta['paper_code'],
                                'Paper Name': current_meta['paper_name']
                            })
                        except Exception as e:
                            st.error(f"❌ Failed to process {file}: {e}")

    if all_rows:
        df_new_sitting_plan = pd.DataFrame(all_rows, columns=sitting_plan_columns)
        
        # Load existing sitting plan if available to merge
        existing_sitting_plan_df = pd.DataFrame()
        if os.path.exists(output_sitting_plan_path):
            try:
                # Ensure correct dtype for roll numbers
                existing_sitting_plan_df = pd.read_csv(output_sitting_plan_path, dtype={f"Roll Number {i}": str for i in range(1, 11)})
                existing_sitting_plan_df.columns = existing_sitting_plan_df.columns.str.strip() # Clean column names
                if 'Paper Code' in existing_sitting_plan_df.columns:
                    existing_sitting_plan_df['Paper Code'] = existing_sitting_plan_df['Paper Code'].apply(_format_paper_code)
            except Exception as e:
                st.warning(f"Could not load existing sitting plan data for update: {e}. Starting fresh for sitting plan.")
                existing_sitting_plan_df = pd.DataFrame(columns=sitting_plan_columns) # Create empty with all cols

        # Align columns before concatenation
        for col in existing_sitting_plan_df.columns:
            if col not in df_new_sitting_plan.columns:
                df_new_sitting_plan[col] = pd.NA
        for col in df_new_sitting_plan.columns:
            if col not in existing_sitting_plan_df.columns:
                existing_sitting_plan_df[col] = pd.NA
        
        df_new_sitting_plan = df_new_sitting_plan[existing_sitting_plan_df.columns] # Reorder new DF columns
        
        combined_sitting_plan_df = pd.concat([existing_sitting_plan_df, df_new_sitting_plan], ignore_index=True)
        
        # Define subset for dropping duplicates in sitting plan (unique student-paper-room assignment)
        roll_num_cols = [f"Roll Number {i+1}" for i in range(10)]
        subset_cols_sitting_plan = roll_num_cols + ["Class", "Mode", "Type", "Room Number", "Paper", "Paper Code", "Paper Name"]
        existing_subset_cols_sitting_plan = [col for col in subset_cols_sitting_plan if col in combined_sitting_plan_df.columns]

        # Fill NA before dropping duplicates to treat them as distinct
        combined_sitting_plan_df_filled = combined_sitting_plan_df.fillna('')
        df_sitting_plan_final = combined_sitting_plan_df_filled.drop_duplicates(subset=existing_subset_cols_sitting_plan, keep='first')
        
        df_sitting_plan_final.to_csv(output_sitting_plan_path, index=False)
        st.success(f"Successfully processed {processed_files_count} PDFs and updated sitting plan to {output_sitting_plan_path}")
    else:
        st.warning("No roll numbers extracted from PDFs to update sitting plan.")

    # Process unique exams for timetable generation
    if unique_exams_for_timetable:
        df_new_timetable_entries = pd.DataFrame(unique_exams_for_timetable).drop_duplicates().reset_index(drop=True)
        
        # Define expected columns for the timetable, including SN, Date, Shift, Time
        expected_columns = ["SN", "Date", "Shift", "Time", "Class", "Paper", "Paper Code", "Paper Name"]
        
        # Load existing timetable if available
        if os.path.exists(output_timetable_path):
            try:
                existing_timetable_df = pd.read_csv(output_timetable_path)
                existing_timetable_df.columns = existing_timetable_df.columns.str.strip() # Clean column names
                if 'Paper Code' in existing_timetable_df.columns:
                    existing_timetable_df['Paper Code'] = existing_timetable_df['Paper Code'].astype(str).str.strip()
            except Exception as e:
                st.warning(f"Could not load existing timetable: {e}. Starting fresh.")
                existing_timetable_df = pd.DataFrame(columns=expected_columns) # Create empty with all cols
        else:
            existing_timetable_df = pd.DataFrame(columns=expected_columns) # Create empty if file doesn't exist

        # Add missing columns to new entries dataframe before concatenating
        for col in expected_columns:
            if col not in df_new_timetable_entries.columns:
                df_new_timetable_entries[col] = pd.NA # Use pd.NA for missing values

            if col not in existing_timetable_df.columns: # Also add to existing if missing
                existing_timetable_df[col] = pd.NA

        # Ensure order of columns for concatenation
        df_new_timetable_entries = df_new_timetable_entries[expected_columns]
        existing_timetable_df = existing_timetable_df[expected_columns] # Reorder existing DF columns

        combined_df = pd.concat([existing_timetable_df, df_new_timetable_entries], ignore_index=True)
        
        # Drop duplicates based on unique exam fields
        unique_fields = ["Date", "Shift", "Time", "Class", "Paper", "Paper Code", "Paper Name"]
        # Ensure 'Time' is included in unique fields for proper deduplication
        # Ensure all unique_fields are in combined_df before dropping duplicates
        unique_fields = [f for f in unique_fields if f in combined_df.columns]


        df_timetable_final = combined_df.drop_duplicates(subset=unique_fields, keep='first').reset_index(drop=True)
        
        # Re-assign SN column sequentially
        df_timetable_final["SN"] = range(1, len(df_timetable_final) + 1)
        
        df_timetable_final.to_csv(output_timetable_path, index=False)
        st.success(f"Timetable updated at {output_timetable_path}.")
        return True, "Timetable deduplicated and saved successfully."
    else:
        st.warning("No unique exam details found to generate timetable.")
        return False, "No data to process."

    return True, "PDF processing complete."

def process_attestation_pdfs(zip_file_buffer, output_csv_path):
    all_data = []

    def parse_pdf_content(text):
        students = re.split(r"\n?RollNo\.\:\s*", text)
        students = [s.strip() for s in students if s.strip()] # Clean empty strings

        student_records = []
        for s in students:
            lines = s.splitlines()
            lines = [line.strip() for line in lines if line.strip()] # Clean each line

            def extract_after(label):
                for i, line in enumerate(lines):
                    if line.startswith(label):
                        value = line.replace(label, "", 1).strip()
                        if value: # Check if value is not empty on the same line
                            return value
                        elif i + 1 < len(lines): # Check next line
                            return lines[i+1].strip()
                    # Special handling for "Regular/Backlog" as it might appear as "Regular/ Backlog:"
                    if label == "Regular/ Backlog:" and line.startswith("Regular/Backlog"):
                        value = line.replace("Regular/Backlog", "", 1).strip()
                        if value:
                            return value
                        elif i + 1 < len(lines):
                            return lines[i+1].strip()
                return "" # Return empty string if not found

            roll_no = re.match(r"(\d{9})", lines[0]).group(1) if lines and re.match(r"(\d{9})", lines[0]) else ""
            enrollment = extract_after("Enrollment No.:")
            session = extract_after("Session:")
            regular = extract_after("Regular/ Backlog:") # Updated label
            student_name = extract_after("Name:")
            father = extract_after("Father's Name:")
            mother = extract_after("Mother's Name:")
            gender = extract_after("Gender:")
            exam_name = extract_after("Exam Name:")
            centre = extract_after("Exam Centre:")
            college = extract_after("College Nmae:") # Typo corrected to 'College Name' if present in PDF
            address = extract_after("Address:")

            papers = re.findall(r"([^\n]+?\[\d{5}\][^\n]*)", s) # Extract papers with [5-digit code]

            student_data = {
                "Roll Number": roll_no,
                "Enrollment Number": enrollment,
                "Session": session,
                "Regular/Backlog": regular,
                "Name": student_name,
                "Father's Name": father,
                "Mother's Name": mother,
                "Gender": gender,
                "Exam Name": exam_name,
                "Exam Centre": centre,
                "College Name": college, # Corrected key
                "Address": address
            }
            # Add up to 10 papers
            for i, paper in enumerate(papers[:10]):
                student_data[f"Paper {i+1}"] = paper.strip()
            
            student_records.append(student_data)
        return student_records

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_file_buffer, 'r') as zip_ref:
            zip_ref.extractall(tmpdir)

        # Determine the base directory for PDFs (e.g., if extracted to 'rasa_pdf/')
        pdf_base_dir = tmpdir
        if 'rasa_pdf' in os.listdir(tmpdir) and os.path.isdir(os.path.join(tmpdir, 'rasa_pdf')):
            pdf_base_dir = os.path.join(tmpdir, 'rasa_pdf')
        elif len(os.listdir(tmpdir)) == 1 and os.path.isdir(os.path.join(tmpdir, os.listdir(tmpdir)[0])):
            pdf_base_dir = os.path.join(tmpdir, os.listdir(tmpdir)[0]) # If single subfolder

        processed_files_count = 0
        for filename in os.listdir(pdf_base_dir):
            if filename.lower().endswith(".pdf"):
                pdf_path = os.path.join(pdf_base_dir, filename)
                try:
                    doc = fitz.open(pdf_path)
                    text = "\n".join([page.get_text() for page in doc])
                    doc.close()
                    st.info(f"📄 Extracting: {filename}")
                    all_data.extend(parse_pdf_content(text))
                    processed_files_count += 1
                except Exception as e:
                    st.error(f"❌ Failed to process {filename}: {e}")

    if all_data:
        df = pd.DataFrame(all_data)
        df.to_csv(output_csv_path, index=False)
        return True, f"Successfully processed {processed_files_count} attestation PDFs and saved to {output_csv_path}"
    else:
        return False, "No data extracted from attestation PDFs."

def generate_college_statistics(input_csv_path, output_csv_path):
    if not os.path.exists(input_csv_path):
        return False, f"Input file not found: {input_csv_path}. Please process attestation PDFs first."
    try:
        df = pd.read_csv(input_csv_path, dtype={"Roll Number": str, "Enrollment Number": str})
        
        # Clean and standardize 'College Name', 'Exam Name', 'Regular/Backlog'
        df['College Name'] = df['College Name'].fillna('UNKNOWN').astype(str).str.strip().str.upper()
        df['Exam Name'] = df['Exam Name'].fillna('UNKNOWN').astype(str).str.strip().str.upper()
        df['Regular/Backlog'] = df['Regular/Backlog'].astype(str).str.strip().str.upper()

        def extract_class_group_and_year(exam_name):
            if pd.isna(exam_name):
                return "UNKNOWN", "UNKNOWN"
            exam_name = str(exam_name).upper().strip()
            # Regex to capture Class Group (e.g., BA, BSC) and Year/Sem
            match = re.match(r'^([A-Z]+)\s*-\s*.+\[\w+\]\s*-\s*(\d+(ST|ND|RD|TH)?(YEAR|SEM))$', exam_name)
            if match:
                class_group = match.group(1).strip()
                year_or_sem = match.group(2).strip()
                return class_group, year_or_sem
            
            # Fallback for Roman numerals (e.g., I, II, III YEAR/SEM)
            roman = re.search(r'\b([IVXLCDM]+)\s*(YEAR|SEM)\b', exam_name)
            if roman:
                return "UNKNOWN", roman.group(0).strip() # Return the full match for year/sem
            
            return "UNKNOWN", "UNKNOWN" # Default if no pattern matches

        df[["Class Group", "Year"]] = df["Exam Name"].apply(lambda x: pd.Series(extract_class_group_and_year(x)))
        
        class_groups = sorted(df["Class Group"].dropna().unique())
        college_list = sorted(df["College Name"].dropna().unique())

        def get_counts(df, college, group, year):
            subset = df[(df["College Name"] == college) & 
                        (df["Class Group"] == group) & 
                        (df["Year"] == year)]
            total = len(subset)
            regular = len(subset[subset["Regular/Backlog"] == "REGULAR"])
            private = len(subset[subset["Regular/Backlog"] == "PRIVATE"])
            exr = len(subset[subset["Regular/Backlog"] == "EXR"])
            supp = len(subset[subset["Regular/Backlog"] == "SUPP"])
            atkt = len(subset[subset["Regular/Backlog"] == "ATKT"])
            return [total, regular, private, exr, atkt, supp]

        output_rows = []

        for group in class_groups:
            years = sorted(df[df["Class Group"] == group]["Year"].dropna().unique())
            # Header for each class group
            header_row1 = ["Class"] + [f"{group} - {year}" for year in years for _ in range(5)]
            header_row2 = ["College", "Grand Total"] + ["Total", "Regular", "Private", "EXR", "ATKT", "SUPP"] * len(years)
            
            block_data = []
            for college in college_list:
                row = [college]
                grand_total = 0
                for year in years:
                    t, r, p, x, a, s = get_counts(df, college, group, year)
                    row += [t, r, p, x, a, s]
                    grand_total += t
                row.insert(1, grand_total) # Insert Grand Total after College Name
                block_data.append(row)
            
            output_rows.append(header_row1)
            output_rows.append(header_row2)
            output_rows += block_data
            output_rows.append([]) # Spacer row

        # Overall College Totals
        output_rows.append(["College", "Total of all"])
        for college in college_list:
            total = len(df[df["College Name"] == college])
            output_rows.append([college, total])

        # Save to CSV without header (since we are creating custom headers)
        pd.DataFrame(output_rows).to_csv(output_csv_path, index=False, header=False)
        
        return True, f"✅ College statistics saved to {output_csv_path}"

    except Exception as e:
        return False, f"❌ Error generating college statistics: {e}"

def generate_sequential_seats(seat_range_str, num_students):
    generated_seats = []
    seat_range_str = seat_range_str.strip().upper() # Normalize input

    if '-' in seat_range_str:
        start_seat_str, end_seat_str = seat_range_str.split('-')
        
        # Case: Alphanumeric (e.g., 1A-60A)
        if re.match(r'^\d+[A-Z]$', start_seat_str) and re.match(r'^\d+[A-Z]$', end_seat_str):
            start_num = int(re.match(r'^(\d+)', start_seat_str).group(1))
            start_char = re.search(r'([A-Z])$', start_seat_str).group(1)
            end_num = int(re.match(r'^(\d+)', end_seat_str).group(1))
            end_char = re.search(r'([A-Z])$', end_seat_str).group(1)

            if start_char != end_char:
                raise ValueError("For alphanumeric seat ranges (e.g., 1A-60A), the alphabet part must be the same.")

            for i in range(start_num, end_num + 1):
                generated_seats.append(f"{i}{start_char}")
        
        # Case: Numeric (e.g., 1-60)
        elif start_seat_str.isdigit() and end_seat_str.isdigit():
            start_num = int(start_seat_str)
            end_num = int(end_seat_str)
            for i in range(start_num, end_num + 1):
                generated_seats.append(str(i))
        else:
            raise ValueError("Invalid seat number range format. Use '1-60' or '1A-60A'.")
    
    # Case: Single seat number (e.g., 1, 1A)
    elif seat_range_str.isdigit() or re.match(r'^\d+[A-Z]$', seat_range_str):
        generated_seats.append(seat_range_str)
    else:
        raise ValueError("Invalid seat number format. Use a single number, '1-60', or '1A-60A'.")
    
    return generated_seats[:num_students] # Return only enough seats for the students

def get_unassigned_students_for_session(date_str, shift, sitting_plan_df, timetable_df):
    unassigned_roll_numbers_details = {}
    
    # Filter timetable for relevant exams on the selected date/shift
    relevant_tt_exams = timetable_df[
        (timetable_df["Date"].astype(str).str.strip() == date_str) & 
        (timetable_df["Shift"].astype(str).str.strip().str.lower() == shift.lower())
    ].copy()

    if relevant_tt_exams.empty:
        return []

    # Create a unique key for exams in timetable
    relevant_tt_exams['exam_key'] = (
        relevant_tt_exams['Class'].astype(str).str.strip().str.lower() + "_" +
        relevant_tt_exams['Paper'].astype(str).str.strip() + "_" +
        relevant_tt_exams['Paper Code'].astype(str).str.strip() + "_" +
        relevant_tt_exams['Paper Name'].astype(str).str.strip()
    )

    for _, sp_row in sitting_plan_df.iterrows():
        # Create a unique key for exams in sitting plan
        sp_exam_key = (
            str(sp_row['Class']).strip().lower() + "_" +
            str(sp_row['Paper']).strip() + "_" +
            str(sp_row['Paper Code']).strip() + "_" +
            str(sp_row['Paper Name']).strip()
        )
        
        # Check if this sitting plan entry is for an exam relevant to the current date/shift
        if sp_exam_key in relevant_tt_exams['exam_key'].values:
            room_assigned = str(sp_row['Room Number']).strip() # Check if room is assigned in sitting plan
            
            for i in range(1, 11): # Iterate through roll number columns
                roll_col = f"Roll Number {i}"
                if roll_col in sp_row and pd.notna(sp_row[roll_col]) and str(sp_row[roll_col]).strip() != '':
                    roll_num = str(sp_row[roll_col]).strip()
                    
                    # If the room number is empty in sitting plan, this student is unassigned (for this paper)
                    if not room_assigned: # If 'Room Number' in sitting plan is blank/empty
                        unassigned_roll_numbers_details[roll_num] = {
                            'Class': str(sp_row['Class']).strip(),
                            'Paper': str(sp_row['Paper']).strip(),
                            'Paper Code': str(sp_row['Paper Code']).strip(),
                            'Paper Name': str(sp_row['Paper Name']).strip()
                        }
    
    # Convert dictionary to sorted list of dictionaries
    sorted_unassigned_list = []
    for roll, details in sorted(unassigned_roll_numbers_details.items()):
        sorted_unassigned_list.append({
            "Roll Number": roll,
            "Class": details['Class'],
            "Paper": details['Paper'],
            "Paper Code": details['Paper Code'],
            "Paper Name": details['Paper Name']
        })
    return sorted_unassigned_list

def get_session_paper_summary(date_str, shift, sitting_plan_df, assigned_seats_df, timetable_df):
    summary_data = []
    
    # Filter timetable for relevant exams on the selected date/shift
    relevant_tt_exams = timetable_df[
        (timetable_df["Date"].astype(str).str.strip() == date_str) & 
        (timetable_df["Shift"].astype(str).str.strip().str.lower() == shift.lower())
    ].copy()

    if relevant_tt_exams.empty:
        return pd.DataFrame(columns=['Paper Name', 'Paper Code', 'Total Expected', 'Assigned', 'Unassigned'])

    # Iterate through unique papers scheduled for the session
    for _, tt_row in relevant_tt_exams.drop_duplicates(subset=['Paper Code', 'Paper Name']).iterrows():
        paper_code = str(tt_row['Paper Code']).strip()
        paper_name = str(tt_row['Paper Name']).strip()

        expected_rolls = set()
        # Find all students expected for this paper from the sitting plan
        paper_sitting_rows = sitting_plan_df[sitting_plan_df['Paper Code'].astype(str).str.strip() == paper_code]
        for _, sp_row in paper_sitting_rows.iterrows():
            for i in range(1, 11): # Check all 10 roll number columns
                roll_col = f"Roll Number {i}"
                if roll_col in sp_row and pd.notna(sp_row[roll_col]) and str(sp_row[roll_col]).strip() != '':
                    expected_rolls.add(str(sp_row[roll_col]).strip())
        
        total_expected_students = len(expected_rolls)

        # Find students already assigned for this paper, date, and shift
        assigned_rolls_for_paper = set(assigned_seats_df[
            (assigned_seats_df["paper code"].astype(str).str.strip() == paper_code) & # Use lowercase column
            (assigned_seats_df["date"] == date_str) & # Use lowercase column
            (assigned_seats_df["shift"] == shift) # Use lowercase column
        ]["roll number"].astype(str).tolist()) # Use lowercase column

        num_assigned_students = len(assigned_rolls_for_paper)
        num_unassigned_students = total_expected_students - num_assigned_students

        summary_data.append({
            'Paper Name': paper_name,
            'Paper Code': paper_code,
            'Total Expected': total_expected_students,
            'Assigned': num_assigned_students,
            'Unassigned': num_unassigned_students
        })
    return pd.DataFrame(summary_data)

def display_room_occupancy_report(sitting_plan_df, assigned_seats_df, timetable_df):
    st.subheader("📊 Room Occupancy Report")
    st.info("View detailed occupancy for each room for a selected date and shift.")

    if sitting_plan_df.empty or timetable_df.empty:
        st.warning("Please upload 'sitting_plan.csv' and 'timetable.csv' via the Admin Panel to generate this report.")
        return

    # Ensure consistent date and shift formats
    timetable_df['Date'] = timetable_df['Date'].astype(str).str.strip()
    timetable_df['Shift'] = timetable_df['Shift'].astype(str).str.strip()
    assigned_seats_df.columns = assigned_seats_df.columns.str.lower() # Standardize assigned seats columns

    report_date_options = sorted(timetable_df["Date"].dropna().unique())
    report_shift_options = sorted(timetable_df["Shift"].dropna().unique())

    if not report_date_options or not report_shift_options:
        st.info("No exam dates or shifts found in the timetable to generate a report.")
        return
    
    selected_report_date = st.selectbox("Select Date", report_date_options, key="room_report_date")
    selected_report_shift = st.selectbox("Select Shift", report_shift_options, key="room_report_shift")

    if st.button("Generate Room Occupancy Report"):
        relevant_tt_exams = timetable_df[(timetable_df["Date"] == selected_report_date) & 
                                         (timetable_df["Shift"].str.lower() == selected_report_shift.lower())]
        if relevant_tt_exams.empty:
            st.info("No exams scheduled for the selected date and shift to generate room occupancy.")
            return

        unique_exams_in_session = relevant_tt_exams[['Class', 'Paper', 'Paper Code', 'Paper Name']].drop_duplicates()
        
        room_occupancy_data = []
        all_rooms_in_sitting_plan = sitting_plan_df['Room Number'].dropna().astype(str).str.strip().unique()

        for room_num in sorted(all_rooms_in_sitting_plan):
            room_sitting_plan_entries = sitting_plan_df[sitting_plan_df['Room Number'].astype(str).str.strip() == room_num]
            
            expected_students_in_room = 0
            # Calculate expected students for the current session and room based on sitting plan
            for _, sp_row in room_sitting_plan_entries.iterrows():
                # Create a standardized key for comparison
                sp_exam_key = (
                    str(sp_row['Class']).strip().lower() + "_" +
                    str(sp_row['Paper']).strip() + "_" +
                    _format_paper_code(sp_row['Paper Code']) + "_" +
                    str(sp_row['Paper Name']).strip().lower()
                )

                is_relevant_exam = False
                for _, ue_row in unique_exams_in_session.iterrows():
                    ue_exam_key = (
                        str(ue_row['Class']).strip().lower() + "_" +
                        str(ue_row['Paper']).strip() + "_" +
                        _format_paper_code(ue_row['Paper Code']) + "_" +
                        str(ue_row['Paper Name']).strip().lower()
                    )
                    if sp_exam_key == ue_exam_key:
                        is_relevant_exam = True
                        break # Found a match, no need to check other unique exams

                if is_relevant_exam:
                    for i in range(1, 11): # Check all roll number columns
                        roll_col = f"Roll Number {i}"
                        if roll_col in sp_row and pd.notna(sp_row[roll_col]) and str(sp_row[roll_col]).strip() != '':
                            expected_students_in_room += 1
            
            # Count actually assigned students for this room, date, and shift
            room_assigned_students_df = assigned_seats_df[
                (assigned_seats_df["room number"].astype(str).str.strip() == room_num) & # Use lowercase
                (assigned_seats_df["date"] == selected_report_date) & # Use lowercase
                (assigned_seats_df["shift"] == selected_report_shift) # Use lowercase
            ]
            assigned_students_in_room = len(room_assigned_students_df)

            assigned_roll_numbers_list = []
            if not room_assigned_students_df.empty:
                def sort_seat_number_for_display(seat):
                    if isinstance(seat, str):
                        if seat.endswith('A'): return (0, int(seat[:-1])) # A comes first
                        elif seat.endswith('B'): return (1, int(seat[:-1])) # B comes second
                        elif seat.isdigit(): return (2, int(seat)) # Numeric comes last
                    return (3, seat) # Fallback for unparseable seats

                room_assigned_students_df['sort_key'] = room_assigned_students_df['seat number'].apply(sort_seat_number_for_display) # Use lowercase
                sorted_room_assigned = room_assigned_students_df.sort_values(by='sort_key').drop(columns=['sort_key'])
                
                for _, assigned_row in sorted_room_assigned.iterrows():
                    assigned_roll_numbers_list.append(f"{assigned_row['roll number']} (Seat: {assigned_row['seat number']}, Paper: {assigned_row['paper code']})") # Use lowercase

            remaining_capacity = expected_students_in_room - assigned_students_in_room
            occupancy_percentage = (assigned_students_in_room / expected_students_in_room * 100) if expected_students_in_room > 0 else 0

            room_occupancy_data.append({
                'Room Number': room_num,
                'Total Expected Students': expected_students_in_room,
                'Assigned Students': assigned_students_in_room,
                'Remaining Capacity': remaining_capacity,
                'Occupancy (%)': f"{occupancy_percentage:.2f}%",
                'Assigned Roll Numbers (Details)': ", ".join(assigned_roll_numbers_list) if assigned_roll_numbers_list else "N/A"
            })
        
        if room_occupancy_data:
            df_occupancy = pd.DataFrame(room_occupancy_data)
            st.dataframe(df_occupancy)
            
            # Download button for CSV
            csv_occupancy = df_occupancy.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="Download Room Occupancy Report as CSV",
                data=csv_occupancy,
                file_name=f"room_occupancy_report_{selected_report_date}_{selected_report_shift}.csv",
                mime="text/csv",
            )
        else:
            st.info("No room occupancy data found for the selected date and shift.")


def generate_room_chart_report(date_str, shift, sitting_plan_df, assigned_seats_df, timetable_df):
    output_string_parts = []

    required_timetable_cols = ["Date", "Shift", "Time", "Class", "Paper Code", "Paper Name"]
    for col in required_timetable_cols:
        if col not in timetable_df.columns:
            return f"Error: Missing essential column '{col}' in timetable.csv. Please ensure the file is correctly formatted."

    # Standardize assigned_seats_df columns to lowercase for consistency
    assigned_seats_df.columns = assigned_seats_df.columns.str.lower()

    required_assigned_seats_cols = ["roll number", "paper code", "paper name", "room number", "seat number", "date", "shift"]
    for col in required_assigned_seats_cols:
        if col not in assigned_seats_df.columns:
            return f"Error: Missing essential column '{col}' in assigned_seats.csv. Please ensure seats are assigned and the file is correctly formatted."

    # Filter timetable for the specific date and shift
    relevant_tt_exams = timetable_df[
        (timetable_df["Date"].astype(str).str.strip() == date_str) & 
        (timetable_df["Shift"].astype(str).str.strip().str.lower() == shift.lower())
    ]
    if relevant_tt_exams.empty:
        return "No exams found for the selected date and shift to generate room chart."

    exam_time = relevant_tt_exams.iloc[0]["Time"].strip() if "Time" in relevant_tt_exams.columns and not relevant_tt_exams.empty else ""

    unique_classes = relevant_tt_exams['Class'].dropna().astype(str).str.strip().unique()
    class_summary_header = ""
    if len(unique_classes) == 1:
        class_summary_header = f"{unique_classes[0]} Examination {datetime.datetime.now().year}"
    elif len(unique_classes) > 1:
        class_summary_header = f"Various Classes Examination {datetime.datetime.now().year}"
    else:
        class_summary_header = f"Examination {datetime.datetime.now().year}"

    # Header for the CSV output (comma-separated for CSV)
    output_string_parts.append(",,,,,,,,,\nजीवाजी विश्वविद्यालय ग्वालियर ,,,,,,,,,\n")
    output_string_parts.append(f"\"परीक्षा केंद्र :- {st.session_state.center_name} (म. प्र.) कोड :- {st.session_state.center_id.upper()} \",,,,,,,,,\n")
    output_string_parts.append(f"{class_summary_header},,,,,,,,,\n")
    output_string_parts.append(f"Date :- ,,{date_str},,Shift :-,{shift},,Time :- ,,\n")

    # Filter assigned students for the specific session
    assigned_students_for_session = assigned_seats_df[
        (assigned_seats_df["date"] == date_str) & 
        (assigned_seats_df["shift"] == shift)
    ].copy()

    if assigned_students_for_session.empty:
        output_string_parts.append("\nNo students assigned seats for this date and shift.")
        return "".join(output_string_parts)

    # Merge with timetable to get Paper Name and Class for assigned students
    # Ensure paper_code columns are same type before merge
    assigned_students_for_session['paper code'] = assigned_students_for_session['paper code'].astype(str)
    timetable_df['Paper Code'] = timetable_df['Paper Code'].astype(str) # This should be the original timetable column

    assigned_students_for_session = pd.merge(
        assigned_students_for_session,
        timetable_df[['Paper Code', 'Paper Name', 'Class']], # Original column names
        left_on='paper code', # Lowercase for assigned_students_for_session
        right_on='Paper Code', # Original for timetable
        how='left',
        suffixes=('', '_tt') # Suffix for timetable columns to avoid conflicts
    )

    # Use the timetable's Paper Name and Class, falling back to assigned_seats_df if necessary
    assigned_students_for_session['paper name'] = assigned_students_for_session['Paper Name_tt'].fillna(assigned_students_for_session['paper name'])
    assigned_students_for_session['class'] = assigned_students_for_session['Class_tt'].fillna(assigned_students_for_session['class'])

    # Define custom sorting key for seat numbers (e.g., 1A, 1B, 1, 2)
    def sort_seat_number_key(seat):
        if isinstance(seat, str):
            match_a = re.match(r'(\d+)A', seat)
            match_b = re.match(r'(\d+)B', seat)
            if match_a:
                return (0, int(match_a.group(1))) # 'A' seats sort first
            elif match_b:
                return (1, int(match_b.group(1))) # 'B' seats sort second
            elif seat.isdigit():
                return (2, int(seat)) # Pure numeric seats sort last
        return (3, seat) # Fallback for unexpected formats

    assigned_students_for_session['sort_key'] = assigned_students_for_session['seat number'].apply(sort_seat_number_key)
    assigned_students_for_session = assigned_students_for_session.sort_values(by=['room number', 'sort_key']).drop(columns=['sort_key'])

    students_by_room = assigned_students_for_session.groupby('room number')

    for room_num, room_data in students_by_room:
        output_string_parts.append(f"\n,,,कक्ष  :-,{room_num}  ,,,,\n")
        
        # Get unique papers for this room and session to display
        unique_papers_in_room = room_data[['class', 'paper code', 'paper name']].drop_duplicates()

        for _, paper_row in unique_papers_in_room.iterrows():
            paper_class = str(paper_row['class']).strip()
            paper_code = str(paper_row['paper code']).strip()
            paper_name = str(paper_row['paper name']).strip()
            
            students_for_this_paper_in_room = room_data[
                (room_data['paper code'].astype(str).str.strip() == paper_code) &
                (room_data['paper name'].astype(str).str.strip() == paper_name)
            ]
            num_students_for_paper = len(students_for_this_paper_in_room)

            output_string_parts.append(
                f"परीक्षा का नाम (Class - mode - Type),,,प्रश्न पत्र (paper- paper code - paper name),,,,उत्तर पुस्तिकाएं (number of students),,\n"
                f",,,,,,,प्राप्त ,प्रयुक्त ,शेष \n" # Hindi for "Received", "Used", "Remaining"
                f"{paper_class} - Regular - Regular,,,{paper_code} - {paper_name}        ,,,,{num_students_for_paper},,\n" # Assuming Regular for mode/type
            )
            output_string_parts.append(",,,,,,,,,\n") # Spacer row

        output_string_parts.append(",,,,,,,,,\n") # Final spacer before total
        output_string_parts.append(f",,,Total,,,,{len(room_data)},,\n") # Total students in the room
        output_string_parts.append(",,,,,,,,,\n") # Spacer row

        # Display roll number list for the room
        output_string_parts.append("roll number - (room number-seat number) - 20 letters of paper name,,,,,,,,,\n")
        current_line_students = []
        for _, student_row in room_data.iterrows():
            roll_num = str(student_row['roll number']).strip()
            room_num_display = str(student_row['room number']).strip()
            seat_num_display = str(student_row['seat number']).strip()
            paper_name_display = str(student_row['paper name']).strip()
            
            truncated_paper_name = paper_name_display[:20] # Truncate paper name for display
            student_entry = f"{roll_num}( कक्ष-{room_num_display}-सीट-{seat_num_display})-{truncated_paper_name}"
            
            current_line_students.append(student_entry)
            
            if len(current_line_students) == 10: # 10 students per line
                output_string_parts.append(",".join(current_line_students) + "\n")
                current_line_students = []
        
        if current_line_students: # Add any remaining students in the last line
            output_string_parts.append(",".join(current_line_students) + "\n")
        
        output_string_parts.append("\n") # Blank line separating rooms

    return "".join(output_string_parts)

def generate_ufm_print_form(ufm_roll_number, attestation_df, assigned_seats_df, timetable_df, report_date, report_shift, report_paper_code, report_paper_name):
    ufm_roll_number = str(ufm_roll_number).strip()

    # Ensure attestation_df 'Roll Number' is string
    attestation_df['Roll Number'] = attestation_df['Roll Number'].astype(str).str.strip()
    student_details = attestation_df[attestation_df['Roll Number'] == ufm_roll_number]

    if student_details.empty:
        return f"Error: Student with Roll Number {ufm_roll_number} not found in attestation data."
    
    student_detail = student_details.iloc[0] # Get the first matching student's details

    # Standardize assigned_seats_df columns to lowercase
    assigned_seats_df.columns = assigned_seats_df.columns.str.lower()

    # Find relevant assigned seat for the specific exam
    relevant_assigned_seat = assigned_seats_df[
        (assigned_seats_df['roll number'].astype(str).str.strip() == ufm_roll_number) &
        (assigned_seats_df['date'].astype(str).str.strip() == report_date) &
        (assigned_seats_df['shift'].astype(str).str.strip() == report_shift) &
        (assigned_seats_df['paper code'].astype(str).str.strip() == _format_paper_code(report_paper_code)) &
        (assigned_seats_df['paper name'].astype(str).str.strip() == report_paper_name)
    ]
    
    exam_room_number = "N/A"
    exam_paper_code = _format_paper_code(report_paper_code)
    exam_paper_name = report_paper_name
    exam_time = "N/A"
    exam_class = "N/A"

    if not relevant_assigned_seat.empty:
        assigned_info = relevant_assigned_seat.iloc[0]
        exam_room_number = str(assigned_info['room number']).strip() # Use lowercase column

        # Also get time and class from timetable based on this specific exam instance
        matching_timetable_entry = timetable_df[
            (timetable_df['Date'].astype(str).str.strip() == report_date) &
            (timetable_df['Shift'].astype(str).str.strip() == report_shift) &
            (timetable_df['Paper Code'].astype(str).str.strip() == _format_paper_code(report_paper_code)) &
            (timetable_df['Paper Name'].astype(str).str.strip() == report_paper_name)
        ]
        if not matching_timetable_entry.empty:
            exam_time = str(matching_timetable_entry.iloc[0]['Time']).strip()
            exam_class = str(matching_timetable_entry.iloc[0]['Class']).strip()
    else:
        # If no assigned seat, try to get exam time/class from timetable directly
        matching_timetable_entry = timetable_df[
            (timetable_df['Date'].astype(str).str.strip() == report_date) &
            (timetable_df['Shift'].astype(str).str.strip() == report_shift) &
            (timetable_df['Paper Code'].astype(str).str.strip() == _format_paper_code(report_paper_code)) &
            (timetable_df['Paper Name'].astype(str).str.strip() == report_paper_name)
        ]
        if not matching_timetable_entry.empty:
            exam_time = str(matching_timetable_entry.iloc[0]['Time']).strip()
            exam_class = str(matching_timetable_entry.iloc[0]['Class']).strip()
        else:
            exam_time = "Not Found in Timetable"
            exam_class = "Not Found in Timetable"


    form_parts = []
    form_parts.append("--- UFM Case Print Form ---")
    form_parts.append("\n**1. Jiwaji University, Gwalior**")
    form_parts.append(f"\n**2. Class:** {exam_class} - {datetime.datetime.now().strftime('%B')}-{datetime.datetime.now().year}-Examination")
    form_parts.append(f"\n**3. Roll Number:** {ufm_roll_number}")
    form_parts.append(f"\n**4. Name of Student:** {student_detail.get('Name', 'N/A')}")
    form_parts.append(f"   **Address:** {student_detail.get('Address', 'N/A')}")
    form_parts.append(f"\n**5. Father's Name:** {student_detail.get('Father\'s Name', 'N/A')}")
    form_parts.append(f"\n**6. College Name:** {student_detail.get('College Name', 'N/A')}")
    form_parts.append(f"\n**7. Exam Center Name:** {st.session_state.center_name} Code: {st.session_state.center_id.upper()}")
    form_parts.append(f"\n**8. Paper Code & Paper Name:** {exam_paper_code} - {exam_paper_name}")
    form_parts.append(f"\n**9. Date:** {report_date}")
    form_parts.append(f"**10. Time:** {report_shift} Shift ({exam_time})")
    form_parts.append(f"\n**11. Time of UFM:** _________________________")
    form_parts.append(f"**12. Name of Book/Material:** _________________________")
    form_parts.append(f"**13. Number of pages/details:** _________________________")
    form_parts.append(f"\n**Room Number (where UFM occurred):** {exam_room_number}")
    form_parts.append("\n\n_________________________")
    form_parts.append("Signature of Invigilator(s)")
    form_parts.append("\n\n_________________________")
    form_parts.append("Signature of Centre Superintendent")
    form_parts.append("\n\n--- End of UFM Case Print Form ---")

    return "\n".join(form_parts)

def display_report_panel():
    st.subheader("📊 Exam Session Reports")
    sitting_plan, timetable, assigned_seats_df, attestation_df = load_data()
    all_reports_df = load_cs_reports_csv()
    room_invigilators_df = load_room_invigilator_assignments()

    if all_reports_df.empty and room_invigilators_df.empty:
        st.info("No Centre Superintendent reports or invigilator assignments available yet for statistics.")
        return

    if sitting_plan.empty:
        st.info("Sitting plan data is required for full report statistics.")

    # Prepare expected students data from sitting plan
    expected_students_data = []
    if not sitting_plan.empty:
        for idx, row in sitting_plan.iterrows():
            expected_count = 0
            for i in range(1, 11): # Check all roll number columns
                if pd.notna(row.get(f"Roll Number {i}")) and str(row.get(f"Roll Number {i}")).strip() != '':
                    expected_count += 1
            expected_students_data.append({
                'Room Number': str(row['Room Number']).strip(),
                'Class': str(row['Class']).strip(),
                'Paper': str(row['Paper']).strip(),
                'Paper Code': _format_paper_code(row['Paper Code']),
                'Paper Name': str(row['Paper Name']).strip(),
                'Mode': str(row.get('Mode', '')).strip(),
                'Type': str(row.get('Type', '')).strip(),
                'expected_students_count': expected_count
            })
    expected_students_df = pd.DataFrame(expected_students_data)

    # Standardize column names for merging (all lowercase)
    all_reports_df['room_num'] = all_reports_df['room_num'].astype(str).str.strip()
    all_reports_df['paper_code'] = all_reports_df['paper_code'].astype(str).str.strip().str.lower()
    all_reports_df['paper_name'] = all_reports_df['paper_name'].astype(str).str.strip().str.lower()
    all_reports_df['class'] = all_reports_df['class'].astype(str).str.strip().str.lower()

    expected_students_df['Room Number'] = expected_students_df['Room Number'].astype(str).str.strip()
    expected_students_df['Paper Code'] = expected_students_df['Paper Code'].astype(str).str.strip().str.lower()
    expected_students_df['Paper Name'] = expected_students_df['Paper Name'].astype(str).str.strip().str.lower()
    expected_students_df['Class'] = expected_students_df['Class'].astype(str).str.strip().str.lower()

    # Merge reports with expected student counts from sitting plan
    merged_reports_df = pd.merge(
        all_reports_df,
        expected_students_df,
        left_on=['room_num', 'paper_code', 'paper_name', 'class'],
        right_on=['Room Number', 'Paper Code', 'Paper Name', 'Class'],
        how='left',
        suffixes=('_report', '_sp')
    )
    merged_reports_df['expected_students_count'] = merged_reports_df['expected_students_count'].fillna(0).astype(int)

    # Merge with room invigilator assignments
    if not room_invigilators_df.empty:
        room_invigilators_df['date'] = room_invigilators_df['date'].astype(str).str.strip()
        room_invigilators_df['shift'] = room_invigilators_df['shift'].astype(str).str.strip().str.lower()
        room_invigilators_df['room_num'] = room_invigilators_df['room_num'].astype(str).str.strip()
        merged_reports_df = pd.merge(
            merged_reports_df,
            room_invigilators_df[['date', 'shift', 'room_num', 'invigilators']],
            on=['date', 'shift', 'room_num'],
            how='left',
            suffixes=('', '_room_inv')
        )
        merged_reports_df['invigilators'] = merged_reports_df['invigilators'].apply(lambda x: x if isinstance(x, list) else [])
    else:
        merged_reports_df['invigilators'] = [[]] * len(merged_reports_df) # Add empty invigilator list if no data

    st.markdown("---")
    st.subheader("Overall Statistics")
    total_reports = len(merged_reports_df)
    unique_sessions = merged_reports_df['report_key'].nunique()
    total_absent = merged_reports_df['absent_roll_numbers'].apply(len).sum()
    total_ufm = merged_reports_df['ufm_roll_numbers'].apply(len).sum()
    
    total_expected_students = expected_students_df['expected_students_count'].sum() # Sum from sitting plan
    total_present_students = total_expected_students - total_absent
    total_answer_sheets_collected = total_present_students - total_ufm

    overall_attendance_percentage = 0
    if total_expected_students > 0:
        overall_attendance_percentage = (total_present_students / total_expected_students) * 100

    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("Total Reports Submitted", total_reports)
    with col2:
        st.metric("Unique Exam Sessions Reported", unique_sessions)
    with col3:
        st.metric("Total Expected Students", total_expected_students)
    with col4:
        st.metric("Total Absent Students", total_absent)
    with col5:
        st.metric("Overall Attendance (%)", f"{overall_attendance_percentage:.2f}%")

    col_metrics_2_1, col_metrics_2_2, col_metrics_2_3 = st.columns(3)
    with col_metrics_2_1:
        st.metric("Total Present Students", total_present_students)
    with col_metrics_2_2:
        st.metric("Total UFM Cases", total_ufm)
    with col_metrics_2_3:
        st.metric("Total Answer Sheets Collected", total_answer_sheets_collected)

    st.markdown("---")
    st.subheader("Paper-wise Statistics")
    # Group by paper name and code from expected_students_df for total expected
    expected_by_paper = expected_students_df.groupby(['Paper Name', 'Paper Code']).agg(
        expected_students=('expected_students_count', 'sum')
    ).reset_index()
    expected_by_paper.rename(columns={'Paper Name': 'paper_name', 'Paper Code': 'paper_code'}, inplace=True)
    expected_by_paper['paper_name'] = expected_by_paper['paper_name'].astype(str).str.strip().str.lower()
    expected_by_paper['paper_code'] = expected_by_paper['paper_code'].astype(str).str.strip().str.lower()

    # Group by paper from reports for absent/UFM
    reported_by_paper = merged_reports_df.groupby(['paper_name', 'paper_code']).agg(
        total_absent=('absent_roll_numbers', lambda x: x.apply(len).sum()),
        total_ufm=('ufm_roll_numbers', lambda x: x.apply(len).sum())
    ).reset_index()

    paper_stats = pd.merge(expected_by_paper, reported_by_paper, on=['paper_name', 'paper_code'], how='left')
    paper_stats['total_absent'] = paper_stats['total_absent'].fillna(0).astype(int)
    paper_stats['total_ufm'] = paper_stats['total_ufm'].fillna(0).astype(int)
    paper_stats['total_present'] = paper_stats['expected_students'] - paper_stats['total_absent']
    paper_stats['total_answer_sheets_collected'] = paper_stats['total_present'] - paper_stats['total_ufm']
    paper_stats['attendance_percentage'] = paper_stats.apply(
        lambda row: (row['total_present'] / row['expected_students'] * 100) if row['expected_students'] > 0 else 0, axis=1
    )
    paper_stats['attendance_percentage'] = paper_stats['attendance_percentage'].map('{:.2f}%'.format)

    paper_stats.rename(columns={
        'paper_name': 'Paper Name', 'paper_code': 'Paper Code',
        'expected_students': 'Expected Students', 'total_absent': 'Absent Students',
        'total_present': 'Present Students', 'total_ufm': 'UFM Cases',
        'total_answer_sheets_collected': 'Answer Sheets Collected',
        'attendance_percentage': 'Attendance (%)'
    }, inplace=True)
    st.dataframe(paper_stats[['Paper Name', 'Paper Code', 'Expected Students', 'Present Students', 'Absent Students', 'UFM Cases', 'Answer Sheets Collected', 'Attendance (%)']])

    st.markdown("---")
    st.subheader("Student Type-wise Statistics")
    expected_by_type = expected_students_df.groupby(['Class', 'Mode', 'Type']).agg(
        expected_students=('expected_students_count', 'sum')
    ).reset_index()
    expected_by_type.rename(columns={'Class': 'Class_sp', 'Mode': 'Mode_sp', 'Type': 'Type_sp'}, inplace=True)

    required_type_cols_for_grouping = ['Class_sp', 'Mode_sp', 'Type_sp']
    
    # Check if these columns exist in merged_reports_df before grouping
    if all(col in merged_reports_df.columns for col in required_type_cols_for_grouping):
        # Create a copy to avoid SettingWithCopyWarning
        reported_by_type_df = merged_reports_df.dropna(subset=required_type_cols_for_grouping).copy()

        if not reported_by_type_df.empty:
            reported_by_type = reported_by_type_df.groupby(required_type_cols_for_grouping).agg(
                total_absent=('absent_roll_numbers', lambda x: x.apply(len).sum()),
                total_ufm=('ufm_roll_numbers', lambda x: x.apply(len).sum())
            ).reset_index()

            type_stats = pd.merge(expected_by_type, reported_by_type, on=required_type_cols_for_grouping, how='left')
            type_stats['total_absent'] = type_stats['total_absent'].fillna(0).astype(int)
            type_stats['total_ufm'] = type_stats['total_ufm'].fillna(0).astype(int)
            type_stats['total_present'] = type_stats['expected_students'] - type_stats['total_absent']
            type_stats['total_answer_sheets_collected'] = type_stats['total_present'] - type_stats['total_ufm']
            type_stats['attendance_percentage'] = type_stats.apply(
                lambda row: (row['total_present'] / row['expected_students'] * 100) if row['expected_students'] > 0 else 0, axis=1
            )
            type_stats['attendance_percentage'] = type_stats['attendance_percentage'].map('{:.2f}%'.format)

            type_stats.rename(columns={
                'Class_sp': 'Class', 'Mode_sp': 'Mode', 'Type_sp': 'Type',
                'expected_students': 'Expected Students', 'total_absent': 'Absent Students',
                'total_present': 'Present Students', 'total_ufm': 'UFM Cases',
                'total_answer_sheets_collected': 'Answer Sheets Collected',
                'attendance_percentage': 'Attendance (%)'
            }, inplace=True)
            st.dataframe(type_stats[['Class', 'Mode', 'Type', 'Expected Students', 'Present Students', 'Absent Students', 'UFM Cases', 'Answer Sheets Collected', 'Attendance (%)']])
        else:
            st.info("No student type data available in reports for statistics after filtering.")
    else:
        st.info("Required student type columns (Class, Mode, Type) are not available in the merged reports for statistics.")


    st.markdown("---")
    st.subheader("Filter and View Reports")
    unique_dates = sorted(merged_reports_df['date'].unique())
    unique_shifts = sorted(merged_reports_df['shift'].unique())
    unique_rooms = sorted(merged_reports_df['room_num'].unique())
    unique_papers = sorted(merged_reports_df['paper_name'].unique())

    filter_date = st.selectbox("Filter by Date", ["All"] + unique_dates, key="report_filter_date")
    filter_shift = st.selectbox("Filter by Shift", ["All"] + unique_shifts, key="report_filter_shift")
    filter_room = st.selectbox("Filter by Room Number", ["All"] + unique_rooms, key="report_filter_room")
    filter_paper = st.selectbox("Filter by Paper Name", ["All"] + unique_papers, key="report_filter_paper")

    filtered_reports_df = merged_reports_df.copy()

    if filter_date != "All":
        filtered_reports_df = filtered_reports_df[filtered_reports_df['date'] == filter_date]
    if filter_shift != "All":
        filtered_reports_df = filtered_reports_df[filtered_reports_df['shift'] == filter_shift]
    if filter_room != "All":
        filtered_reports_df = filtered_reports_df[filtered_reports_df['room_num'] == filter_room]
    if filter_paper != "All":
        filtered_reports_df = filtered_reports_df[filtered_reports_df['paper_name'] == filter_paper]

    if filtered_reports_df.empty:
        st.info("No reports match the selected filters.")
    else:
        st.markdown("---")
        st.subheader("Filtered Reports Summary")
        st.dataframe(filtered_reports_df[['date', 'shift', 'room_num', 'paper_code', 'paper_name', 'invigilators', 'absent_roll_numbers', 'ufm_roll_numbers']].rename(columns={
            'date': 'Date', 'shift': 'Shift', 'room_num': 'Room', 'paper_code': 'Paper Code', 
            'paper_name': 'Paper Name', 'invigilators': 'Invigilators',
            'absent_roll_numbers': 'Absent Roll Numbers', 'ufm_roll_numbers': 'UFM Roll Numbers'
        }))

        st.markdown("---")
        st.subheader("Detailed Absentee List (Filtered)")
        absent_list_data = []
        for _, row in filtered_reports_df.iterrows():
            for roll in row['absent_roll_numbers']:
                absent_list_data.append({
                    'Date': row['date'], 'Shift': row['shift'], 'Room': row['room_num'],
                    'Paper Code': row['paper_code'], 'Paper Name': row['paper_name'], 'Absent Roll Number': roll
                })
        if absent_list_data:
            df_absent = pd.DataFrame(absent_list_data)
            st.dataframe(df_absent)
            csv_absent = df_absent.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="Download Absentee List as CSV",
                data=csv_absent,
                file_name=f"absent_list_{filter_date}_{filter_shift}.csv",
                mime="text/csv",
            )
        else:
            st.info("No absent students in the filtered reports.")

        st.markdown("---")
        st.subheader("Detailed UFM List (Filtered)")
        ufm_list_data = []
        for _, row in filtered_reports_df.iterrows():
            for roll in row['ufm_roll_numbers']:
                ufm_list_data.append({
                    'Date': row['date'], 'Shift': row['shift'], 'Room': row['room_num'],
                    'Paper Code': row['paper_code'], 'Paper Name': row['paper_name'], 'UFM Roll Number': roll
                })
        if ufm_list_data:
            df_ufm = pd.DataFrame(ufm_list_data)
            st.dataframe(df_ufm)
            csv_ufm = df_ufm.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="Download UFM List as CSV",
                data=csv_ufm,
                file_name=f"ufm_list_{filter_date}_{filter_shift}.csv",
                mime="text/csv",
            )
        else:
            st.info("No UFM cases in the filtered reports.")


def calculate_remuneration(shift_assignments_df, room_invigilator_assignments_df, timetable_df, assigned_seats_df, manual_rates, prep_closing_assignments, holiday_dates, selected_classes_for_bill):
    remuneration_data_detailed_raw = []

    # Define remuneration rules and rates
    remuneration_rules = {
        'senior_center_superintendent': {'role_display': 'Senior Center Superintendent', 'rate': manual_rates['senior_center_superintendent_rate'], 'unit': 'day', 'eligible_prep_close': True, 'exam_conveyance': True},
        'center_superintendent': {'role_display': 'Center Superintendent', 'rate': manual_rates['center_superintendent_rate'], 'unit': 'shift', 'eligible_prep_close': True, 'exam_conveyance': True},
        'assistant_center_superintendent': {'role_display': 'Assistant Center Superintendent', 'rate': manual_rates['assistant_center_superintendent_rate'], 'unit': 'shift', 'eligible_prep_close': True, 'exam_conveyance': True},
        'permanent_invigilator': {'role_display': 'Permanent Invigilator', 'rate': manual_rates['permanent_invigilator_rate'], 'unit': 'shift', 'eligible_prep_close': True, 'exam_conveyance': True},
        'assistant_permanent_invigilator': {'role_display': 'Assistant Permanent Invigilator', 'rate': manual_rates['assistant_permanent_invigilator_rate'], 'unit': 'shift', 'eligible_prep_close': False, 'exam_conveyance': True},
        'invigilator': {'role_display': 'Invigilator', 'rate': manual_rates['invigilator_rate'], 'unit': 'shift', 'eligible_prep_close': False, 'exam_conveyance': True},
    }

    class_worker_rates = {
        'class_3_worker': {'role_display': 'Class 3 Worker', 'rate_per_student': manual_rates['class_3_worker_rate_per_student']},
        'class_4_worker': {'role_display': 'Class 4 Worker', 'rate_per_student': manual_rates['class_4_worker_rate_per_student']},
    }

    unified_assignments = []
    unique_class_3_workers = set()
    unique_class_4_workers = set()

    # Process shift assignments
    for index, row in shift_assignments_df.iterrows():
        current_date = row['date'] # Ensure 'date' column is correctly accessed (now lowercase)
        current_shift = row['shift'] # Ensure 'shift' column is correctly accessed (now lowercase)
        for role_col in remuneration_rules.keys():
            if role_col in row and isinstance(row[role_col], list):
                for person in row[role_col]:
                    unified_assignments.append({
                        'Name': person,
                        'Role_Key': role_col,
                        'Date': current_date,
                        'Shift': current_shift,
                        'Source': 'shift_assignments'
                    })
        if 'class_3_worker' in row and isinstance(row['class_3_worker'], list):
            unique_class_3_workers.update(row['class_3_worker'])
        if 'class_4_worker' in row and isinstance(row['class_4_worker'], list):
            unique_class_4_workers.update(row['class_4_worker'])

    # Process room invigilator assignments
    for index, row in room_invigilator_assignments_df.iterrows():
        current_date = row['date'] # Ensure 'date' column is correctly accessed (now lowercase)
        current_shift = row['shift'] # Ensure 'shift' column is correctly accessed (now lowercase)
        invigilators_list = row['invigilators'] # This should be a list due to parsing
        for invigilator in invigilators_list:
            # Check if this invigilator is already assigned a higher role for this date/shift
            is_assigned_higher_role = False
            for assignment in unified_assignments:
                if (assignment['Name'] == invigilator and 
                    assignment['Date'] == current_date and 
                    assignment['Shift'] == current_shift and 
                    assignment['Role_Key'] != 'invigilator'):
                    is_assigned_higher_role = True
                    break
            
            if not is_assigned_higher_role:
                unified_assignments.append({
                    'Name': invigilator,
                    'Role_Key': 'invigilator',
                    'Date': current_date,
                    'Shift': current_shift,
                    'Source': 'room_invigilator_assignments'
                })

    df_assignments = pd.DataFrame(unified_assignments)

    # Determine classes per session for filtering
    session_classes_map = {}
    if not timetable_df.empty:
        for _, tt_row in timetable_df.iterrows():
            date_shift_key = (str(tt_row['Date']).strip(), str(tt_row['Shift']).strip())
            if date_shift_key not in session_classes_map:
                session_classes_map[date_shift_key] = set()
            session_classes_map[date_shift_key].add(str(tt_row['Class']).strip())

    # Identify workers who worked both shifts on a day for conveyance calculation
    workers_with_both_shifts = set()
    if not df_assignments.empty:
        # Convert date column to datetime objects for accurate grouping/sorting
        df_assignments['Date_dt'] = pd.to_datetime(df_assignments['Date'], format='%d-%m-%Y', errors='coerce')
        # Count unique shifts per person per day
        shift_counts = df_assignments.groupby(['Name', 'Date_dt'])['Shift'].nunique().reset_index()
        eligible_workers_df = shift_counts[shift_counts['Shift'] == 2] # People with 2 unique shifts
        for _, row in eligible_workers_df.iterrows():
            workers_with_both_shifts.add((row['Name'], row['Date_dt'].strftime('%d-%m-%Y'))) # Store as string for consistency

    # Calculate detailed remuneration per person per shift
    for assignment in unified_assignments:
        name = assignment['Name']
        role_key = assignment['Role_Key']
        date = assignment['Date']
        shift = assignment['Shift']
        
        session_classes = list(session_classes_map.get((date, shift), set()))
        
        # Check if this assignment is for a selected class exam
        is_selected_exam = any(cls in selected_classes_for_bill for cls in [c.strip() for c in session_classes]) if selected_classes_for_bill else True

        base_rem_for_shift = remuneration_rules[role_key]['rate']
        conveyance = 0

        if remuneration_rules[role_key]['exam_conveyance']:
            if (name, date) in workers_with_both_shifts:
                # Conveyance is applied to the evening shift only if they worked both
                if shift == 'Evening' and is_selected_exam:
                    conveyance = manual_rates['conveyance_rate']
                elif shift == 'Morning' and is_selected_exam: # Only morning shift, no conveyance
                    conveyance = 0

        remuneration_data_detailed_raw.append({
            'Name': name,
            'Role_Key': role_key,
            'Role_Display': remuneration_rules[role_key]['role_display'],
            'Date': date,
            'Shift': shift,
            'Base_Remuneration_Per_Shift_Unfiltered': base_rem_for_shift,
            'Conveyance': conveyance,
            'Is_Selected_Exam': is_selected_exam,
            'Classes_in_Session': session_classes,
        })

    df_detailed_remuneration = pd.DataFrame(remuneration_data_detailed_raw)

    # --- Generate Individual Bills ---
    individual_bills = []
    unique_person_roles = df_detailed_remuneration[['Name', 'Role_Display', 'Role_Key']].drop_duplicates()

    for idx, row in unique_person_roles.iterrows():
        name = row['Name']
        role_display = row['Role_Display']
        role_key = row['Role_Key']

        person_data = df_detailed_remuneration[(df_detailed_remuneration['Name'] == name) & 
                                               (df_detailed_remuneration['Role_Display'] == role_display)].copy()

        # Special handling for Senior Center Superintendent (remuneration per day, not per shift if both shifts worked)
        if role_key == 'senior_center_superintendent':
            # Filter for shifts linked to selected classes
            selected_shifts = person_data[person_data['Is_Selected_Exam'] == True]
            non_selected_shifts = person_data[person_data['Is_Selected_Exam'] == False]

            selected_dates = {} # date -> list of shifts (Morning/Evening)
            non_selected_dates = {}

            for _, shift_row in selected_shifts.iterrows():
                date = shift_row['Date']
                selected_dates.setdefault(date, []).append(shift_row['Shift'])
            
            for _, shift_row in non_selected_shifts.iterrows():
                date = shift_row['Date']
                non_selected_dates.setdefault(date, []).append(shift_row['Shift'])

            # Determine actual eligible shifts for SCS
            eligible_shifts = []
            for _, row in person_data.iterrows():
                date = row['Date']
                shift = row['Shift']
                is_selected = row['Is_Selected_Exam']
                
                # If morning shift and is selected exam, count it
                if shift == 'Morning' and is_selected:
                    eligible_shifts.append(row)
                elif shift == 'Evening':
                    if is_selected:
                        # If evening selected, and morning was NOT selected for this date, count evening.
                        # This prevents double counting for full days for SCS for selected exams
                        if 'Morning' not in selected_dates.get(date, []):
                            eligible_shifts.append(row)
                    else:
                        # If evening non-selected, and morning was NOT selected for this date, count evening.
                        if 'Morning' not in selected_dates.get(date, []):
                            eligible_shifts.append(row)


            filtered_person_data = pd.DataFrame(eligible_shifts)
        else:
            # For all other roles, filter by Is_Selected_Exam if classes are selected
            if selected_classes_for_bill:
                filtered_person_data = person_data[person_data['Is_Selected_Exam'] == True].copy()
            else:
                filtered_person_data = person_data.copy()


        # Format duty dates for display
        duty_dates_morning_str = ""
        morning_shifts_df = filtered_person_data[filtered_person_data['Shift'] == 'Morning']
        if not morning_shifts_df.empty:
            morning_shifts_df['Date_dt'] = pd.to_datetime(morning_shifts_df['Date'], format='%d-%m-%Y', errors='coerce')
            morning_shifts_df = morning_shifts_df.sort_values(by='Date_dt')
            grouped_dates_morning = morning_shifts_df.groupby(morning_shifts_df['Date_dt'].dt.to_period('M'))['Date_dt'].apply(lambda x: sorted(x.dt.day.tolist()))
            date_parts = []
            for period, days in grouped_dates_morning.items():
                month_name = period.strftime('%b')
                days_str = ", ".join(map(str, days))
                date_parts.append(f"{month_name} - {days_str}")
            if date_parts:
                duty_dates_morning_str = ", ".join(date_parts) + f" {morning_shifts_df['Date_dt'].min().year}"

        duty_dates_evening_str = ""
        evening_shifts_df = filtered_person_data[filtered_person_data['Shift'] == 'Evening']
        if not evening_shifts_df.empty:
            evening_shifts_df['Date_dt'] = pd.to_datetime(evening_shifts_df['Date'], format='%d-%m-%Y', errors='coerce')
            evening_shifts_df = evening_shifts_df.sort_values(by='Date_dt')
            grouped_dates_evening = evening_shifts_df.groupby(evening_shifts_df['Date_dt'].dt.to_period('M'))['Date_dt'].apply(lambda x: sorted(x.dt.day.tolist()))
            date_parts = []
            for period, days in grouped_dates_evening.items():
                month_name = period.strftime('%b')
                days_str = ", ".join(map(str, days))
                date_parts.append(f"{month_name} - {days_str}")
            if date_parts:
                duty_dates_evening_str = ", ".join(date_parts) + f" {evening_shifts_df['Date_dt'].min().year}"

        total_morning_shifts = len(morning_shifts_df)
        total_evening_shifts = len(evening_shifts_df)
        total_shifts = total_morning_shifts + total_evening_shifts

        rate_in_rs = remuneration_rules[role_key]['rate'] if role_key in remuneration_rules else 0

        total_base_remuneration = 0
        if role_key == 'senior_center_superintendent':
            # SCS is paid per day, not per shift. Count unique dates from filtered data.
            unique_dates = filtered_person_data['Date'].nunique()
            total_base_remuneration = unique_dates * rate_in_rs
        else:
            # Other roles are paid per shift
            total_base_remuneration = filtered_person_data['Base_Remuneration_Per_Shift_Unfiltered'].sum()
        
        # Total Conveyance (only from Evening shifts if worked both)
        total_conveyance = person_data['Conveyance'].sum()

        # Preparation and Closing day remuneration
        total_prep_remuneration = 0
        total_closing_remuneration = 0
        total_holiday_conveyance = 0

        if remuneration_rules[role_key]['eligible_prep_close']:
            person_assignments = prep_closing_assignments.get(name, {})
            assigned_role = person_assignments.get('role')

            # Only add prep/closing remuneration if the assigned role matches the current bill's role
            if assigned_role == role_key: 
                prep_days = person_assignments.get('prep_days', [])
                total_prep_remuneration = len(prep_days) * rate_in_rs

                closing_days = person_assignments.get('closing_days', [])
                total_closing_remuneration = len(closing_days) * rate_in_rs

                # Holiday conveyance for prep/closing days
                all_assigned_days = prep_days + closing_days
                holiday_assigned_days = [day for day in all_assigned_days if day in holiday_dates]
                total_holiday_conveyance = len(holiday_assigned_days) * manual_rates['holiday_conveyance_allowance_rate']

        grand_total_amount = total_base_remuneration + total_conveyance + total_prep_remuneration + total_closing_remuneration + total_holiday_conveyance
        
        individual_bills.append({
            'SN': len(individual_bills) + 1,
            'Name (with role)': f"{name} ({role_display})",
            'Duty dates of selected class exam Shift (morning)': duty_dates_morning_str,
            'Duty dates of selected class exam Shift (evening)': duty_dates_evening_str,
            'Total shifts of selected class exams (morning/evening)': total_shifts,
            'Rate in Rs': rate_in_rs,
            'Total Remuneration in Rs': total_base_remuneration,
            'Total Conveyance (in evening shift)': total_conveyance,
            'Preparation Day Remuneration': total_prep_remuneration,
            'Closing Day Remuneration': total_closing_remuneration,
            'Total Holiday Conveyance Added': total_holiday_conveyance,
            'Total amount in Rs': grand_total_amount,
            'Signature': ''
        })

    df_individual_bills = pd.DataFrame(individual_bills)

    # --- Generate Role Summary Matrix ---
    df_role_summary_matrix = generate_role_summary_matrix_by_date(df_detailed_remuneration, remuneration_rules, prep_closing_assignments, holiday_dates, manual_rates, selected_classes_for_bill)

    # --- Generate Class 3 and Class 4 Worker Bills ---
    class_3_4_final_bills = []

    # Calculate total students based on selected classes for bill or all students
    if selected_classes_for_bill:
        # Get paper codes for selected classes from timetable
        papers_for_selected_classes = timetable_df[timetable_df['Class'].isin(selected_classes_for_bill)]['Paper Code'].unique()
        # Filter assigned seats by these paper codes
        # Ensure 'paper code' is lowercase in assigned_seats_df for consistency
        assigned_seats_df.columns = assigned_seats_df.columns.str.lower()
        filtered_assigned_seats = assigned_seats_df[assigned_seats_df['paper code'].isin(papers_for_selected_classes)]
        total_students_for_class_workers = filtered_assigned_seats['roll number'].nunique() # Use lowercase
    else:
        # If no classes selected, count all unique students assigned
        # Ensure 'roll number' is lowercase
        assigned_seats_df.columns = assigned_seats_df.columns.str.lower()
        total_students_for_class_workers = assigned_seats_df['roll number'].nunique()

    if unique_class_3_workers:
        class_3_total_fixed_amount = total_students_for_class_workers * class_worker_rates['class_3_worker']['rate_per_student']
        num_class_3_workers = len(unique_class_3_workers)
        rem_per_class_3_worker = class_3_total_fixed_amount / num_class_3_workers if num_class_3_workers > 0 else 0
        for sn, worker_name in enumerate(sorted(list(unique_class_3_workers))):
            class_3_4_final_bills.append({
                'S.N.': len(class_3_4_final_bills) + 1,
                'Name': worker_name,
                'Role': class_worker_rates['class_3_worker']['role_display'],
                'Total Students (Center-wide)': total_students_for_class_workers,
                'Rate per Student (for category)': class_worker_rates['class_3_worker']['rate_per_student'],
                'Total Remuneration for Category (Rs.)': class_3_total_fixed_amount,
                'Number of Workers in Category': num_class_3_workers,
                'Remuneration per Worker in Rs.': rem_per_class_3_worker,
                'Signature of Receiver': ''
            })

    if unique_class_4_workers:
        class_4_total_fixed_amount = total_students_for_class_workers * class_worker_rates['class_4_worker']['rate_per_student']
        num_class_4_workers = len(unique_class_4_workers)
        rem_per_class_4_worker = class_4_total_fixed_amount / num_class_4_workers if num_class_4_workers > 0 else 0
        for sn, worker_name in enumerate(sorted(list(unique_class_4_workers))):
            class_3_4_final_bills.append({
                'S.N.': len(class_3_4_final_bills) + 1,
                'Name': worker_name,
                'Role': class_worker_rates['class_4_worker']['role_display'],
                'Total Students (Center-wide)': total_students_for_class_workers,
                'Rate per Student (for category)': class_4_total_fixed_amount,
                'Number of Workers in Category': num_class_4_workers,
                'Remuneration per Worker in Rs.': rem_per_class_4_worker,
                'Signature of Receiver': ''
            })
    df_class_3_4_final_bills = pd.DataFrame(class_3_4_final_bills)

    return df_individual_bills, df_role_summary_matrix, df_class_3_4_final_bills

def generate_role_summary_matrix_by_date(df_detailed_remuneration, remuneration_rules, prep_closing_assignments, holiday_dates, manual_rates, selected_classes_for_bill):
    # Collect all unique dates from exam assignments and prep/closing assignments
    prep_closing_dates = set()
    for name, assignments in prep_closing_assignments.items():
        prep_closing_dates.update(assignments.get('prep_days', []))
        prep_closing_dates.update(assignments.get('closing_days', []))
    
    exam_dates = set()
    # Filter exam dates to only include those with activity for selected classes or conveyance
    for date_str in df_detailed_remuneration['Date'].unique():
        date_data = df_detailed_remuneration[df_detailed_remuneration['Date'] == date_str]
        has_work = False
        for _, person_assignment in date_data.iterrows():
            if person_assignment['Is_Selected_Exam'] or person_assignment['Conveyance'] > 0:
                has_work = True
                break
        if has_work:
            exam_dates.add(date_str)


    all_dates = exam_dates.union(prep_closing_dates)
    
    # Sort dates chronologically
    sorted_dates = sorted([pd.to_datetime(d, format='%d-%m-%Y', errors='coerce') for d in all_dates if pd.notna(pd.to_datetime(d, format='%d-%m-%Y', errors='coerce'))])

    columns = [
        'Date & Shift', 'senior_center_superintendent', 'center_superintendent',
        'assistant_center_superintendent', 'permanent_invigilator',
        'assistant_permanent_invigilator', 'invigilator', 'Conveyance', 'Daily Total'
    ]
    matrix_data = []

    # Iterate through each date to populate the matrix
    for date_dt in sorted_dates:
        date_str = date_dt.strftime('%d-%m-%Y')

        # Only process exam dates here, prep/closing dates handled separately below
        if date_str not in exam_dates:
            continue
        
        date_assignments = df_detailed_remuneration[df_detailed_remuneration['Date'] == date_str]
        if date_assignments.empty:
            continue # Skip if no assignments for this date

        # Morning Shift
        row_morning = {'Date & Shift': f"{date_str} (Morning)"}
        for col in columns[1:]: row_morning[col] = 0 # Initialize with zeros

        morning_data = df_detailed_remuneration[(df_detailed_remuneration['Date'] == date_str) & 
                                                (df_detailed_remuneration['Shift'] == 'Morning')]
        morning_has_activity = False # Flag to check if morning shift has any relevant activity

        for _, person_assignment in morning_data.iterrows():
            role_key = person_assignment['Role_Key']
            rem_rate = remuneration_rules[role_key]['rate']
            
            # Only count remuneration if it's for a selected exam
            if person_assignment['Is_Selected_Exam']:
                morning_has_activity = True
                # SCS remuneration is handled per day, not per shift, so exclude here
                if role_key != 'senior_center_superintendent':
                    row_morning[role_key] += rem_rate
        
        # Evening Shift
        row_evening = {'Date & Shift': f"{date_str} (Evening)"}
        for col in columns[1:]: row_evening[col] = 0 # Initialize with zeros
        
        evening_data = df_detailed_remuneration[(df_detailed_remuneration['Date'] == date_str) & 
                                                (df_detailed_remuneration['Shift'] == 'Evening')]
        evening_has_activity = False # Flag to check if evening shift has any relevant activity

        for _, person_assignment in evening_data.iterrows():
            role_key = person_assignment['Role_Key']
            rem_rate = remuneration_rules[role_key]['rate']
            
            # Count if it's a selected exam or if there's conveyance
            if person_assignment['Is_Selected_Exam'] or person_assignment['Conveyance'] > 0:
                evening_has_activity = True
                if role_key != 'senior_center_superintendent':
                    row_evening[role_key] += rem_rate
                row_evening['Conveyance'] += person_assignment['Conveyance']

        # Add Senior Center Superintendent remuneration (calculated per day)
        senior_cs_data = df_detailed_remuneration[(df_detailed_remuneration['Date'] == date_str) & 
                                                  (df_detailed_remuneration['Role_Key'] == 'senior_center_superintendent')]
        if not senior_cs_data.empty:
            name = senior_cs_data['Name'].iloc[0] # Assume one SCS per day for simplicity
            person_data_for_day = df_detailed_remuneration[df_detailed_remuneration['Name'] == name]

            # Re-evaluate SCS eligibility for this specific day (similar logic as in calculate_remuneration)
            selected_shifts = person_data_for_day[person_data_for_day['Is_Selected_Exam'] == True]
            non_selected_shifts = person_data_for_day[person_data_for_day['Is_Selected_Exam'] == False]
            
            selected_dates = set(selected_shifts['Date'])
            non_selected_dates = set(non_selected_shifts['Date']) # Dates where SCS worked non-selected exams

            eligible_for_day = False
            if date_str in selected_dates:
                eligible_for_day = True
                # If SCS worked morning of a selected exam and evening of a non-selected exam, and evening of selected was NOT worked
                # this logic helps ensure SCS gets paid for the whole day (Morning shift amount).
                # This is complex and might need adjustment based on exact SCS payment rules.
                if ('Morning' in selected_shifts[selected_shifts['Date'] == date_str]['Shift'].tolist() and
                    'Evening' in non_selected_shifts[non_selected_shifts['Date'] == date_str]['Shift'].tolist() and
                    'Evening' not in selected_shifts[selected_shifts['Date'] == date_str]['Shift'].tolist()):
                    eligible_for_day = False # Prevent double counting if already counted for morning selected

            if eligible_for_day:
                row_morning['senior_center_superintendent'] += remuneration_rules['senior_center_superintendent']['rate']
                morning_has_activity = True # Mark activity for the morning row due to SCS

        row_morning['Daily Total'] = sum(row_morning[col] for col in columns[1:-1])
        row_evening['Daily Total'] = sum(row_evening[col] for col in columns[1:-1])

        # Append rows only if there was relevant activity or total amount is > 0
        if morning_has_activity or row_morning['Daily Total'] > 0:
            matrix_data.append(row_morning)
        if evening_has_activity or row_evening['Daily Total'] > 0:
            matrix_data.append(row_evening)

    # Add Preparation/Closing day remuneration
    prep_closing_remuneration = {} # Stores total remuneration for prep/closing per date per role
    for name, assignments in prep_closing_assignments.items():
        role_key = assignments.get('role')
        if role_key in remuneration_rules: # Ensure role is valid
            rate = remuneration_rules[role_key]['rate']
            
            prep_days = assignments.get('prep_days', [])
            closing_days = assignments.get('closing_days', [])

            for date in prep_days:
                if date not in prep_closing_remuneration:
                    prep_closing_remuneration[date] = {}
                if role_key not in prep_closing_remuneration[date]:
                    prep_closing_remuneration[date][role_key] = 0
                prep_closing_remuneration[date][role_key] += rate # Add remuneration for prep day
            
            for date in closing_days:
                if date not in prep_closing_remuneration:
                    prep_closing_remuneration[date] = {}
                if role_key not in prep_closing_remuneration[date]:
                    prep_closing_remuneration[date][role_key] = 0
                prep_closing_remuneration[date][role_key] += rate # Add remuneration for closing day

    # Append prep/closing rows to the matrix
    for date, payments in prep_closing_remuneration.items():
        row = {'Date & Shift': f"{date} (Prep/Closing)"}
        for col in columns[1:]: row[col] = 0 # Initialize with zeros

        for role_key, payment in payments.items():
            if role_key in row: # Ensure the role column exists
                row[role_key] += payment
            else:
                # Handle cases where a prep/closing role might not be a direct column in the matrix
                # For example, if 'PI/API' is a combined column, it needs custom handling
                pass 
        
        # Add holiday conveyance for prep/closing days
        if date in holiday_dates:
            # Count unique workers assigned to prep/closing on this holiday
            num_workers_on_holiday = sum(1 for person_assignments in prep_closing_assignments.values() 
                                         if date in person_assignments.get('prep_days', []) or 
                                            date in person_assignments.get('closing_days', []))
            row['Conveyance'] += num_workers_on_holiday * manual_rates['holiday_conveyance_allowance_rate']
        
        row['Daily Total'] = sum(row[col] for col in columns[1:-1]) # Sum all values except Date & Shift and Conveyance
        matrix_data.append(row)


    df_matrix = pd.DataFrame(matrix_data)
    
    # Add a 'Total' row at the end
    total_row = {'Date & Shift': 'Total'}
    for col in columns[1:]:
        total_row[col] = df_matrix[col].sum()
    df_matrix = pd.concat([df_matrix, pd.DataFrame([total_row])], ignore_index=True)

    # Combine Permanent Invigilator and Assistant Permanent Invigilator columns if both exist
    # This ensures PI/API column sums up both roles.
    if 'permanent_invigilator' in df_matrix.columns and 'assistant_permanent_invigilator' in df_matrix.columns:
        df_matrix['PI/API'] = df_matrix['permanent_invigilator'] + df_matrix['assistant_permanent_invigilator']
        df_matrix = df_matrix.drop(columns=['permanent_invigilator', 'assistant_permanent_invigilator'])
    else:
        # If one is missing, ensure PI/API is still created, possibly with just one of them or zeros
        if 'permanent_invigilator' in df_matrix.columns:
            df_matrix['PI/API'] = df_matrix['permanent_invigilator']
            df_matrix = df_matrix.drop(columns=['permanent_invigilator'])
        elif 'assistant_permanent_invigilator' in df_matrix.columns:
            df_matrix['PI/API'] = df_matrix['assistant_permanent_invigilator']
            df_matrix = df_matrix.drop(columns=['assistant_permanent_invigilator'])
        else:
            df_matrix['PI/API'] = 0 # If neither exists

    # Final column order and renaming for display
    final_cols = ['Date & Shift', 'SCS', 'CS', 'ACS', 'PI/API', 'Invigilators', 'Conveyance', 'Daily Total']
    df_matrix = df_matrix.rename(columns={
        'senior_center_superintendent': 'SCS',
        'center_superintendent': 'CS',
        'assistant_center_superintendent': 'ACS',
        'invigilator': 'Invigilators' # The 'invigilator' from `remuneration_rules`
    })

    # Ensure all final columns exist, filling with 0 if not present
    for col in final_cols:
        if col not in df_matrix.columns:
            df_matrix[col] = 0

    return df_matrix[final_cols]


def add_total_row(df):
    """Adds a total row to a DataFrame, summing numeric columns."""
    if df.empty:
        return df
    
    total_row = {}
    for col in df.columns:
        if col in ['SN', 'S.N.']: # Identify serial number columns
            total_row[col] = 'TOTAL'
        # Identify columns that should not be summed (e.g., names, dates, signatures)
        elif col in ['Name (with role)', 'Name', 'Role', 'Duty dates', 'Shift (morning/evening)', 'Signature', 
                      'Signature of Receiver', 'Duty dates of selected class exam Shift (morning)',
                      'Duty dates of selected class exam Shift (evening)', 'Date & Shift']:
            total_row[col] = ''
        elif pd.api.types.is_numeric_dtype(df[col]): # Check if column is numeric
            total_row[col] = df[col].sum()
        else:
            total_row[col] = '' # Leave other columns blank
    
    total_df = pd.DataFrame([total_row])
    return pd.concat([df, total_df], ignore_index=True)


def save_bills_to_excel(individual_bills_df, role_summary_df, class_workers_df, filename="remuneration_bills.xlsx"):
    """Saves the remuneration bills to an Excel file with proper formatting."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not individual_bills_df.empty:
            individual_bills_df.to_excel(writer, sheet_name='Individual Bills', index=False)
            worksheet = writer.sheets['Individual Bills']
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter # Get the column letter (e.g., 'A', 'B')
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) # Add a little extra padding
                worksheet.column_dimensions[column_name].width = adjusted_width

        if not role_summary_df.empty:
            role_summary_df.to_excel(writer, sheet_name='Role Summary', index=False)
            worksheet = writer.sheets['Role Summary']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_name].width = adjusted_width

        if not class_workers_df.empty:
            class_workers_df.to_excel(writer, sheet_name='Class Workers', index=False)
            worksheet = writer.sheets['Class Workers']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_name].width = adjusted_width
    
    output.seek(0) # Rewind the buffer to the beginning
    return output, filename

# --- Main Streamlit Application UI ---
st.title("Examination Management System")

# Select Exam Center (already defined above the functions)
center_name = st.selectbox(
    "Select Exam Center",
    list(EXAM_CENTERS.keys()),
    key="selected_center_name"
)
st.session_state.center_name = center_name # Store display name
st.session_state.center_id = EXAM_CENTERS[center_name] # Store unique ID

menu = st.sidebar.radio("Select Module", ["Student View", "Admin Panel", "Centre Superintendent Panel"])

if menu == "Student View":
    sitting_plan, timetable, assigned_seats_df, attestation_df = load_data()
    
    if sitting_plan.empty or timetable.empty:
        st.warning("Sitting plan or timetable data not found. Please upload them via the Admin Panel for full functionality.")

    option = st.radio("Choose Search Option:", ["Search by Roll Number and Date", "Get Full Exam Schedule by Roll Number", "View Full Timetable"])

    if option == "Search by Roll Number and Date":
        roll = st.text_input("Enter Roll Number", max_chars=9)
        date_input = st.date_input("Enter Exam Date", value=datetime.date.today())
        
        if st.button("Search"):
            if sitting_plan.empty or timetable.empty:
                st.warning("Sitting plan or timetable data is missing. Please upload them via the Admin Panel to search.")
            else:
                results = get_sitting_details(roll, date_input.strftime('%d-%m-%Y'), sitting_plan, timetable)
                if results:
                    st.success(f"Found {len(results)} exam(s) for Roll Number {roll} on {date_input.strftime('%d-%m-%Y')}:")
                    for i, result in enumerate(results):
                        st.markdown(f"---")
                        st.subheader(f"Exam {i+1}")
                        st.write(f"**Room Number:** {result['Room Number']}")
                        st.write(f"**🪑 Seat Number:** {result['Seat Number']}")
                        st.write(f"**📚 Paper:** {result['Paper']} - {result['Paper Name']} - ({result['Paper Code']})")
                        st.write(f"**🏫 Class:** {result['Class']}")
                        st.write(f"**🎓 Student type:** {result['Mode']} - {result['Type']}")
                        st.write(f"**🕐 Shift:** {result['Shift']}, **📅 Date:** {result['Date']}")
                else:
                    st.warning("No data found for the given inputs.")

    elif option == "Get Full Exam Schedule by Roll Number":
        roll = st.text_input("Enter Roll Number")
        if st.button("Get Schedule"):
            if sitting_plan.empty or timetable.empty:
                st.warning("Sitting plan or timetable data is missing. Please upload them via the Admin Panel to get schedule.")
            else:
                schedule = pd.DataFrame(get_all_exams(roll, sitting_plan, timetable))
                if not schedule.empty:
                    # Convert 'Date' to datetime for proper sorting
                    schedule['Date_dt'] = pd.to_datetime(schedule['Date'], format='%d-%m-%Y', errors='coerce')
                    schedule = schedule.sort_values(by="Date_dt").drop(columns=['Date_dt'])
                    st.write(schedule)
                else:
                    st.warning("No exam records found for this roll number.")

    elif option == "View Full Timetable":
        st.subheader("Full Examination Timetable")
        if timetable.empty:
            st.warning("Timetable data is missing. Please upload it via the Admin Panel.")
        else:
            st.dataframe(timetable)

elif menu == "Admin Panel":
    st.subheader("🔐 Admin Login")
    if admin_login():
        st.success("Login successful!")
        
        # Load data for the selected center
        sitting_plan, timetable, assigned_seats_df, attestation_df = load_data()

        st.markdown("---")
        st.subheader("Current Data Previews (for selected center)")
        col_sp, col_tt, col_assigned, col_attestation = st.columns(4)
        with col_sp:
            st.write(f"**{SITTING_PLAN_FILE}**")
            if not sitting_plan.empty:
                st.dataframe(sitting_plan)
            else:
                st.info("No sitting plan data loaded.")
        with col_tt:
            st.write(f"**{TIMETABLE_FILE}**")
            if not timetable.empty:
                st.dataframe(timetable)
            else:
                st.info("No timetable data loaded.")
        with col_assigned:
            st.write(f"**{ASSIGNED_SEATS_FILE}**")
            if not assigned_seats_df.empty:
                st.dataframe(assigned_seats_df)
            else:
                st.info("No assigned seats data loaded.")
        with col_attestation:
            st.write(f"**{ATTESTATION_DATA_FILE}**")
            if not attestation_df.empty:
                st.dataframe(attestation_df)
            else:
                st.info("No attestation data loaded.")

        st.markdown("---")
        admin_option = st.radio("Select Admin Task:", [
            "Upload Data Files",
            "Get All Students for Date & Shift (Room Wise)",
            "Get All Students for Date & Shift (Roll Number Wise)",
            "Update Timetable Details",
            "Assign Rooms & Seats to Students",
            "Room Occupancy Report",
            "Room Chart Report",
            "Generate College Statistics",
            "Remuneration Bill Generation",
            "Report Panel",
            "Download Attestation Data",
            "Stop (Reset and Re-upload All CSVs)"
        ])
        
        if admin_option == "Stop (Reset and Re-upload All CSVs)":
            st.subheader("🚨 Reset and Re-upload All Data (for current center)")
            st.warning(f"This will DELETE ALL DATA from Supabase for **'{st.session_state.center_name}'** and re-upload the local CSVs (Sitting Plan, Timetable, Exam Team Members, Shift Assignments, Room Invigilator Assignments, Assigned Seats). Use with caution!")
            confirm_reset = st.checkbox("I understand and want to proceed with deleting and re-uploading data for this center.")
            if confirm_reset and st.button("🔴 Confirm Reset and Re-upload"):
                with st.spinner("Deleting and re-uploading data... This may take a moment."):
                    # Tables to reset (ensure these match your Supabase table names)
                    tables_to_reset = ["sitting_plan", "timetable", "exam_team_members", 
                                       "shift_assignments", "room_invigilator_assignments", 
                                       "assigned_seats", "cs_reports", "attestation_data_combined"] 

                    for table_name in tables_to_reset:
                        st.info(f"Deleting data from '{table_name}' for center '{st.session_state.center_id}'...")
                        try:
                            # Delete all records for the current center_id
                            supabase.table(table_name).delete().eq('center_id', st.session_state.center_id).execute()
                            st.success(f"Deleted existing data from '{table_name}' for this center.")
                        except Exception as e:
                            st.error(f"Error deleting data from '{table_name}': {e}")

                    st.info("Starting re-upload of local CSVs...")

                    upload_success_count = 0
                    upload_fail_count = 0

                    # Re-upload Sitting Plan
                    sitting_plan_csv = get_center_filepath(SITTING_PLAN_FILE)
                    if os.path.exists(sitting_plan_csv):
                        success, msg = upload_csv_to_supabase("sitting_plan", sitting_plan_csv)
                        if success: st.success(msg); upload_success_count += 1
                        else: st.error(msg); upload_fail_count += 1
                    else:
                        st.warning(f"Local {SITTING_PLAN_FILE} not found to re-upload.")

                    # Re-upload Timetable
                    timetable_csv = get_center_filepath(TIMETABLE_FILE)
                    if os.path.exists(timetable_csv):
                        success, msg = upload_csv_to_supabase("timetable", timetable_csv)
                        if success: st.success(msg); upload_success_count += 1
                        else: st.error(msg); upload_fail_count += 1
                    else:
                        st.warning(f"Local {TIMETABLE_FILE} not found to re-upload.")

                    # Re-upload Exam Team Members
                    exam_team_csv = get_center_filepath(EXAM_TEAM_MEMBERS_FILE)
                    if os.path.exists(exam_team_csv):
                        success, msg = upload_csv_to_supabase("exam_team_members", exam_team_csv)
                        if success: st.success(msg); upload_success_count += 1
                        else: st.error(msg); upload_fail_count += 1
                    else:
                        st.warning(f"Local {EXAM_TEAM_MEMBERS_FILE} not found to re-upload.")
                    
                    # Re-upload Shift Assignments
                    shift_assignments_csv = get_center_filepath(SHIFT_ASSIGNMENTS_FILE)
                    if os.path.exists(shift_assignments_csv):
                        success, msg = upload_csv_to_supabase("shift_assignments", shift_assignments_csv)
                        if success: st.success(msg); upload_success_count += 1
                        else: st.error(msg); upload_fail_count += 1
                    else:
                        st.warning(f"Local {SHIFT_ASSIGNMENTS_FILE} not found to re-upload.")

                    # Re-upload Room Invigilator Assignments
                    room_invigilators_csv = get_center_filepath(ROOM_INVIGILATORS_FILE)
                    if os.path.exists(room_invigilators_csv):
                        success, msg = upload_csv_to_supabase("room_invigilator_assignments", room_invigilators_csv)
                        if success: st.success(msg); upload_success_count += 1
                        else: st.error(msg); upload_fail_count += 1
                    else:
                        st.warning(f"Local {ROOM_INVIGILATORS_FILE} not found to re-upload.")

                    # Re-upload Assigned Seats (if any have been saved locally)
                    assigned_seats_csv = get_center_filepath(ASSIGNED_SEATS_FILE)
                    if os.path.exists(assigned_seats_csv):
                        success, msg = upload_csv_to_supabase("assigned_seats", assigned_seats_csv)
                        if success: st.success(msg); upload_success_count += 1
                        else: st.error(msg); upload_fail_count += 1
                    else:
                        st.warning(f"Local {ASSIGNED_SEATS_FILE} not found to re-upload.")

                    # Re-upload Attestation Data
                    attestation_data_csv = get_center_filepath(ATTESTATION_DATA_FILE)
                    if os.path.exists(attestation_data_csv):
                        success, msg = upload_csv_to_supabase("attestation_data_combined", attestation_data_csv)
                        if success: st.success(msg); upload_success_count += 1
                        else: st.error(msg); upload_fail_count += 1
                    else:
                        st.warning(f"Local {ATTESTATION_DATA_FILE} not found to re-upload.")

                    st.success(f"Re-upload process complete. {upload_success_count} files re-uploaded, {upload_fail_count} failed.")
                    st.rerun() # Rerun to refresh data previews

        elif admin_option == "Download Attestation Data":
            st.subheader("⬇️ Download Attestation Data")
            st.info(f"Click the button below to download '{ATTESTATION_DATA_FILE}' from Supabase to the current center's data folder.")
            if st.button("Download Attestation Data"):
                download_attestation_data_to_center_folder()

        elif admin_option == "Upload Data Files":
            st.subheader("⬆️ Upload Data Files (for selected center)")
            st.info("Upload your essential CSV files here. They will be saved locally for the selected center and uploaded to Supabase.")
            
            # Upload Sitting Plan
            uploaded_sitting_plan = st.file_uploader(f"Upload {SITTING_PLAN_FILE}", type=["csv"], key="upload_sitting_plan")
            if uploaded_sitting_plan:
                success, msg = save_uploaded_file(uploaded_sitting_plan, SITTING_PLAN_FILE)
                if success:
                    st.success(msg)
                    # Also upload to Supabase
                    supabase_success, supabase_msg = upload_csv_to_supabase("sitting_plan", get_center_filepath(SITTING_PLAN_FILE))
                    if supabase_success:
                        st.success(f"Supabase: {supabase_msg}")
                        st.rerun() # Rerun to load updated data
                    else:
                        st.error(f"Supabase Upload Error: {supabase_msg}")
                else:
                    st.error(msg)

            # Upload Timetable
            uploaded_timetable = st.file_uploader(f"Upload {TIMETABLE_FILE}", type=["csv"], key="upload_timetable")
            if uploaded_timetable:
                success, msg = save_uploaded_file(uploaded_timetable, TIMETABLE_FILE)
                if success:
                    st.success(msg)
                    supabase_success, supabase_msg = upload_csv_to_supabase("timetable", get_center_filepath(TIMETABLE_FILE))
                    if supabase_success:
                        st.success(f"Supabase: {supabase_msg}")
                        st.rerun()
                    else:
                        st.error(f"Supabase Upload Error: {supabase_msg}")
                else:
                    st.error(msg)
            
            # Upload Exam Team Members
            uploaded_exam_team = st.file_uploader(f"Upload {EXAM_TEAM_MEMBERS_FILE}", type=["csv"], key="upload_exam_team")
            if uploaded_exam_team:
                success, msg = save_uploaded_file(uploaded_exam_team, EXAM_TEAM_MEMBERS_FILE)
                if success:
                    st.success(msg)
                    supabase_success, supabase_msg = upload_csv_to_supabase("exam_team_members", get_center_filepath(EXAM_TEAM_MEMBERS_FILE))
                    if supabase_success:
                        st.success(f"Supabase: {supabase_msg}")
                        st.rerun()
                    else:
                        st.error(f"Supabase Upload Error: {supabase_msg}")
                else:
                    st.error(msg)

            # Upload Shift Assignments
            uploaded_shift_assignments = st.file_uploader(f"Upload {SHIFT_ASSIGNMENTS_FILE}", type=["csv"], key="upload_shift_assignments")
            if uploaded_shift_assignments:
                success, msg = save_uploaded_file(uploaded_shift_assignments, SHIFT_ASSIGNMENTS_FILE)
                if success:
                    st.success(msg)
                    supabase_success, supabase_msg = upload_csv_to_supabase("shift_assignments", get_center_filepath(SHIFT_ASSIGNMENTS_FILE))
                    if supabase_success:
                        st.success(f"Supabase: {supabase_msg}")
                        st.rerun()
                    else:
                        st.error(f"Supabase Upload Error: {supabase_msg}")
                else:
                    st.error(msg)

            # Upload Room Invigilator Assignments
            uploaded_room_invigilators = st.file_uploader(f"Upload {ROOM_INVIGILATORS_FILE}", type=["csv"], key="upload_room_invigilators")
            if uploaded_room_invigilators:
                success, msg = save_uploaded_file(uploaded_room_invigilators, ROOM_INVIGILATORS_FILE)
                if success:
                    st.success(msg)
                    supabase_success, supabase_msg = upload_csv_to_supabase("room_invigilator_assignments", get_center_filepath(ROOM_INVIGILATORS_FILE))
                    if supabase_success:
                        st.success(f"Supabase: {supabase_msg}")
                        st.rerun()
                    else:
                        st.error(f"Supabase Upload Error: {supabase_msg}")
                else:
                    st.error(msg)

            st.markdown("---")
            st.subheader("Process & Upload Attestation PDFs")
            st.info("Upload a ZIP file containing attestation PDFs (e.g., 'rasa_pdf.zip'). This will extract data and create 'attestation_data_combined.csv'.")
            uploaded_attestation_zip = st.file_uploader("Upload Attestation PDFs (ZIP)", type=["zip"], key="upload_attestation_zip")
            if uploaded_attestation_zip:
                # Use center-specific path for the output CSV
                output_attestation_csv_path = get_center_filepath(ATTESTATION_DATA_FILE)
                with st.spinner("Processing Attestation PDFs..."):
                    success, msg = process_attestation_pdfs(uploaded_attestation_zip, output_attestation_csv_path)
                    if success:
                        st.success(msg)
                        # Immediately upload to Supabase after processing
                        supabase_success, supabase_msg = upload_csv_to_supabase("attestation_data_combined", output_attestation_csv_path)
                        if supabase_success:
                            st.success(f"Supabase: {supabase_msg}")
                            st.rerun()
                        else:
                            st.error(f"Supabase Upload Error: {supabase_msg}")
                    else:
                        st.error(msg)

            st.markdown("---")
            st.subheader("Process & Upload Sitting Plan PDFs")
            st.info("Upload a ZIP file containing sitting plan PDFs. This will extract data and create/update 'sitting_plan.csv' and 'timetable.csv'.")
            uploaded_sitting_plan_zip = st.file_uploader("Upload Sitting Plan PDFs (ZIP)", type=["zip"], key="upload_sitting_plan_zip")
            if uploaded_sitting_plan_zip:
                # Use center-specific paths for output CSVs
                output_sitting_plan_csv_path = get_center_filepath(SITTING_PLAN_FILE)
                output_timetable_csv_path = get_center_filepath(TIMETABLE_FILE)
                with st.spinner("Processing Sitting Plan PDFs..."):
                    success, msg = process_sitting_plan_pdfs(uploaded_sitting_plan_zip, output_sitting_plan_csv_path, output_timetable_csv_path)
                    if success:
                        st.success(msg)
                        # Upload updated sitting plan and timetable to Supabase
                        supabase_sp_success, supabase_sp_msg = upload_csv_to_supabase("sitting_plan", output_sitting_plan_csv_path)
                        if supabase_sp_success:
                            st.success(f"Supabase: {supabase_sp_msg}")
                        else:
                            st.error(f"Supabase Upload Error (Sitting Plan): {supabase_sp_msg}")
                        
                        supabase_tt_success, supabase_tt_msg = upload_csv_to_supabase("timetable", output_timetable_csv_path)
                        if supabase_tt_success:
                            st.success(f"Supabase: {supabase_tt_msg}")
                        else:
                            st.error(f"Supabase Upload Error (Timetable): {supabase_tt_msg}")
                        
                        if supabase_sp_success and supabase_tt_success:
                            st.rerun() # Rerun to load updated data
                    else:
                        st.error(msg)
        
        elif admin_option == "Get All Students for Date & Shift (Room Wise)":
            st.subheader("List All Students for a Date and Shift (Room Wise)")
            if assigned_seats_df.empty or timetable.empty:
                st.info("Please ensure seats are assigned and 'timetable.csv' is uploaded to use this feature.")
            else:
                list_date_input = st.date_input("Select Date", value=datetime.date.today())
                list_shift_options = ["Morning", "Evening"]
                list_shift = st.selectbox("Select Shift", list_shift_options)

                if st.button("Get Student List (Room Wise)"):
                    formatted_student_list_text, error_message, excel_data_for_students_list = get_all_students_for_date_shift_formatted(list_date_input.strftime('%d-%m-%Y'), list_shift, assigned_seats_df, timetable)
                    if formatted_student_list_text:
                        st.success(f"Generated list for {list_date_input.strftime('%d-%m-%Y')} ({list_shift} Shift):")
                        st.text_area("Student List (Text Format)", formatted_student_list_text, height=500)
                        
                        file_name_txt = (f"all_students_list_room_wise_{list_date_input.strftime('%Y%m%d')}_"
                                         f"{list_shift.lower()}.txt")
                        st.download_button(
                            label="Download Student List (Room Wise) as TXT",
                            data=formatted_student_list_text,
                            file_name=file_name_txt,
                            mime="text/plain"
                        )
                        
                        if excel_data_for_students_list:
                            output = io.BytesIO()
                            workbook = Workbook()
                            sheet = workbook.active
                            sheet.title = "Student List (Room Wise)"
                            for row_data in excel_data_for_students_list:
                                sheet.append(row_data)
                            
                            # Adjust column widths in Excel
                            for col_idx, col_cells in enumerate(sheet.columns):
                                max_length = 0
                                for cell in col_cells:
                                    try:
                                        if cell.value is not None:
                                            cell_value_str = str(cell.value)
                                            # Handle multi-line cells for width calculation
                                            current_length = max(len(line) for line in cell_value_str.split('\n'))
                                            if current_length > max_length:
                                                max_length = current_length
                                    except Exception as e:
                                        st.error(f"Error processing cell: {e}")
                                        pass # Skip if error occurs
                                
                                adjusted_width = (max_length + 2) # Add a little padding
                                sheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
                            
                            workbook.save(output)
                            processed_data = output.getvalue()
                            
                            file_name_excel = (f"all_students_list_room_wise_{list_date_input.strftime('%Y%m%d')}_"
                                               f"{list_shift.lower()}.xlsx")
                            st.download_button(
                                label="Download Student List (Room Wise) as Excel",
                                data=processed_data,
                                file_name=file_name_excel,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.warning(f"No students found: {error_message}")
        
        elif admin_option == "Get All Students for Date & Shift (Roll Number Wise)":
            st.subheader("List All Students for a Date and Shift (Roll Number Wise)")
            if assigned_seats_df.empty or timetable.empty:
                st.info("Please ensure seats are assigned and 'timetable.csv' is uploaded to use this feature.")
            else:
                list_date_input = st.date_input("Select Date", value=datetime.date.today(), key="roll_num_wise_date")
                list_shift_options = ["Morning", "Evening"]
                list_shift = st.selectbox("Select Shift", list_shift_options, key="roll_num_wise_shift")

                if st.button("Get Student List (Roll Number Wise)"):
                    formatted_student_list_text, error_message, excel_data_for_students_list = get_all_students_roll_number_wise_formatted(list_date_input.strftime('%d-%m-%Y'), list_shift, assigned_seats_df, timetable)
                    if formatted_student_list_text:
                        st.success(f"Generated list for {list_date_input.strftime('%d-%m-%Y')} ({list_shift} Shift):")
                        st.text_area("Student List (Text Format)", formatted_student_list_text, height=500)
                        
                        file_name_txt = (f"all_students_list_roll_wise_{list_date_input.strftime('%Y%m%d')}_"
                                         f"{list_shift.lower()}.txt")
                        st.download_button(
                            label="Download Student List (Roll Number Wise) as TXT",
                            data=formatted_student_list_text,
                            file_name=file_name_txt,
                            mime="text/plain"
                        )
                        
                        if excel_data_for_students_list:
                            output = io.BytesIO()
                            workbook = Workbook()
                            sheet = workbook.active
                            sheet.title = "Student List (Roll Wise)"
                            for row_data in excel_data_for_students_list:
                                sheet.append(row_data)
                            
                            # Adjust column widths in Excel
                            for col_idx, col_cells in enumerate(sheet.columns):
                                max_length = 0
                                for cell in col_cells:
                                    try:
                                        if cell.value is not None:
                                            cell_value_str = str(cell.value)
                                            current_length = max(len(line) for line in cell_value_str.split('\n'))
                                            if current_length > max_length:
                                                max_length = current_length
                                    except Exception as e:
                                        st.error(f"Error processing cell: {e}")
                                        pass # Skip if error occurs
                                
                                adjusted_width = (max_length + 2) # Add a little padding
                                sheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
                            
                            workbook.save(output)
                            processed_data = output.getvalue()
                            
                            file_name_excel = (f"all_students_list_roll_wise_{list_date_input.strftime('%Y%m%d')}_"
                                               f"{list_shift.lower()}.xlsx")
                            st.download_button(
                                label="Download Student List (Roll Number Wise) as Excel",
                                data=processed_data,
                                file_name=file_name_excel,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.warning(f"No students found: {error_message}")
        
        elif admin_option == "Update Timetable Details":
            st.subheader("✏️ Update Timetable Details")
            if timetable.empty:
                st.info("No timetable data loaded. Please upload 'timetable.csv' first using the 'Upload Data Files' section.")
            else:
                st.write("Current Timetable Preview:")
                st.dataframe(timetable)
                st.markdown("---")

                st.write("Select filters to specify which entries to update:")
                unique_dates_tt = sorted(timetable['Date'].astype(str).unique().tolist())
                unique_shifts_tt = sorted(timetable['Shift'].astype(str).unique().tolist())
                unique_classes_tt = sorted(timetable['Class'].astype(str).unique().tolist())
                unique_paper_codes_tt = sorted(timetable['Paper Code'].astype(str).unique().tolist())
                unique_paper_tt = sorted(timetable['Paper'].astype(str).unique().tolist())
                unique_paper_names_tt = sorted(timetable['Paper Name'].astype(str).unique().tolist())

                filter_date_tt_update = st.selectbox("Filter by Date", ["All"] + unique_dates_tt, key="filter_date_tt_update")
                filter_shift_tt_update = st.selectbox("Filter by Shift", ["All"] + unique_shifts_tt, key="filter_shift_tt_update")
                filter_class_tt_update = st.selectbox("Filter by Class", ["All"] + unique_classes_tt, key="filter_class_tt_update")
                filter_paper_code_tt_update = st.selectbox("Filter by Paper Code", ["All"] + unique_paper_codes_tt, key="filter_paper_code_tt_update")
                filter_paper_tt_update = st.selectbox("Filter by Paper", ["All"] + unique_paper_tt, key="filter_paper_tt_update")
                filter_paper_name_tt_update = st.selectbox("Filter by Paper Name", ["All"] + unique_paper_names_tt, key="filter_paper_name_tt_update")

                st.markdown("---")
                st.write("Entries that will be updated based on your filters:")
                temp_filtered_tt = timetable.copy()
                if filter_date_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Date'].astype(str) == filter_date_tt_update]
                if filter_shift_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Shift'].astype(str) == filter_shift_tt_update]
                if filter_class_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Class'].astype(str) == filter_class_tt_update]
                if filter_paper_code_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Paper Code'].astype(str) == filter_paper_code_tt_update]
                if filter_paper_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Paper'].astype(str) == filter_paper_tt_update]
                if filter_paper_name_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Paper Name'].astype(str) == filter_paper_name_tt_update]

                if temp_filtered_tt.empty:
                    st.info("No entries match the selected filters. No updates will be applied.")
                else:
                    st.dataframe(temp_filtered_tt)

                st.markdown("---")
                st.write("Enter new values for 'Date', 'Shift', and 'Time' for the filtered entries:")
                
                # Set default values for update inputs
                default_date_update_input = datetime.date.today()
                if not temp_filtered_tt.empty and 'Date' in temp_filtered_tt.columns and pd.notna(temp_filtered_tt['Date'].iloc[0]):
                    try:
                        default_date_update_input = datetime.datetime.strptime(str(temp_filtered_tt['Date'].iloc[0]).strip(), '%d-%m-%Y').date()
                    except ValueError:
                        pass # Keep default if parsing fails
                elif 'Date' in timetable.columns and not timetable['Date'].empty and pd.notna(timetable['Date'].iloc[0]):
                     try: # Fallback to first date in full timetable
                        default_date_update_input = datetime.datetime.strptime(str(timetable['Date'].iloc[0]).strip(), '%d-%m-%Y').date()
                     except ValueError:
                        pass # Keep default if parsing fails

                default_shift_update_input = "Morning"
                if not temp_filtered_tt.empty and 'Shift' in temp_filtered_tt.columns and pd.notna(temp_filtered_tt['Shift'].iloc[0]):
                    default_shift_update_input = str(temp_filtered_tt['Shift'].iloc[0]).strip()
                elif 'Shift' in timetable.columns and not timetable['Shift'].empty and pd.notna(timetable['Shift'].iloc[0]):
                    default_shift_update_input = str(timetable['Shift'].iloc[0]).strip()


                default_time_update_input = "09:00 AM - 12:00 PM"
                if not temp_filtered_tt.empty and 'Time' in temp_filtered_tt.columns and pd.notna(temp_filtered_tt['Time'].iloc[0]):
                    default_time_update_input = str(temp_filtered_tt['Time'].iloc[0]).strip()
                elif 'Time' in timetable.columns and not timetable['Time'].empty and pd.notna(timetable['Time'].iloc[0]):
                    default_time_update_input = str(timetable['Time'].iloc[0]).strip()

                update_date = st.date_input("New Date", value=default_date_update_input, key="update_tt_date")
                update_shift = st.selectbox("New Shift", ["Morning", "Evening"], 
                                            index=["Morning", "Evening"].index(default_shift_update_input) if default_shift_update_input in ["Morning", "Evening"] else 0,
                                            key="update_tt_shift")
                update_time = st.text_input("New Time (e.g., 09:00 AM - 12:00 PM)", value=default_time_update_input, key="update_tt_time")

                if st.button("Apply Updates and Save Timetable"):
                    if temp_filtered_tt.empty:
                        st.warning("No entries matched your filters, so no updates were applied.")
                    else:
                        timetable_modified = timetable.copy()
                        # Get indices of rows to update using the original timetable and filters
                        indices_to_update = timetable_modified[
                            (timetable_modified['Date'].astype(str) == filter_date_tt_update if filter_date_tt_update != "All" else True) &
                            (timetable_modified['Shift'].astype(str) == filter_shift_tt_update if filter_shift_tt_update != "All" else True) &
                            (timetable_modified['Class'].astype(str) == filter_class_tt_update if filter_class_tt_update != "All" else True) &
                            (timetable_modified['Paper Code'].astype(str) == filter_paper_code_tt_update if filter_paper_code_tt_update != "All" else True) &
                            (timetable_modified['Paper'].astype(str) == filter_paper_tt_update if filter_paper_tt_update != "All" else True) &
                            (timetable_modified['Paper Name'].astype(str) == filter_paper_name_tt_update if filter_paper_name_tt_update != "All" else True)
                        ].index

                        if not indices_to_update.empty:
                            timetable_modified.loc[indices_to_update, 'Date'] = update_date.strftime('%d-%m-%Y')
                            timetable_modified.loc[indices_to_update, 'Shift'] = update_shift
                            timetable_modified.loc[indices_to_update, 'Time'] = update_time
                            
                            success, msg = save_uploaded_file(timetable_modified, get_center_filepath(TIMETABLE_FILE))
                            if success:
                                st.success(f"Timetable details updated for {len(indices_to_update)} entries and saved successfully.")
                                # Upload to Supabase after saving locally
                                supabase_success, supabase_msg = upload_csv_to_supabase("timetable", get_center_filepath(TIMETABLE_FILE))
                                if supabase_success:
                                    st.success(f"Supabase: {supabase_msg}")
                                    load_data() # Reload data to reflect changes in UI
                                    st.rerun() # Rerun to refresh Streamlit state
                                else:
                                    st.error(f"Supabase Upload Error: {supabase_msg}")
                            else:
                                st.error(msg)
                        else:
                            st.warning("No entries matched your filters to apply updates.")
        
        elif admin_option == "Assign Rooms & Seats to Students":
            st.subheader("📘 Room & Seat Assignment Tool")
            st.markdown("""This tool helps manage seat assignments for exams, offering real-time status updates,
                            capacity warnings, and clear error messages based on your selected seat format.""")

            if sitting_plan.empty or timetable.empty:
                st.error(f"Error: `{SITTING_PLAN_FILE}` or `{TIMETABLE_FILE}` not found. Please upload these files to run the assignment tool.")
                st.stop() # Stop execution if essential data is missing

            # Initialize session state variables for room status
            if 'current_room_status_a_rem' not in st.session_state:
                st.session_state.current_room_status_a_rem = None
            if 'current_room_status_b_rem' not in st.session_state:
                st.session_state.current_room_status_b_rem = None
            if 'current_room_status_total_rem' not in st.session_state:
                st.session_state.current_room_status_total_rem = None

            st.subheader("Exam Details")
            date_options = sorted(timetable["Date"].dropna().unique())
            shift_options = sorted(timetable["Shift"].dropna().unique())

            if not date_options or not shift_options:
                st.warning("Timetable is empty or missing Date/Shift information. Please upload a complete timetable.")
            else:
                date = st.selectbox("Select Exam Date", date_options, key="assign_date_select")
                shift = st.selectbox("Select Shift", shift_options, key="assign_shift_select")
                
                st.markdown("---")
                st.subheader("Session Student Summary (Assigned vs. Unassigned)")
                session_paper_summary_df = get_session_paper_summary(date, shift, sitting_plan, assigned_seats_df, timetable)
                if not session_paper_summary_df.empty:
                    st.dataframe(session_paper_summary_df)
                else:
                    st.info("No student data found for the selected date and shift.")
                
                st.markdown("---")
                filtered_papers = timetable[(timetable["Date"] == date) & (timetable["Shift"] == shift)]
                paper_options = filtered_papers[["Paper Code", "Paper Name"]].drop_duplicates().values.tolist()
                paper_display = [f"{_format_paper_code(code)} - {name}" for code, name in paper_options]
                selected_paper = st.selectbox("Select Paper Code and Name", paper_display, key="assign_paper_select")

                if selected_paper:
                    paper_code = _format_paper_code(selected_paper.split(" - ")[0])
                    paper_name = selected_paper.split(" - ", 1)[1].strip()

                    st.subheader("Room & Seat Configuration")
                    room = st.text_input("Enter Room Number (e.g., 1, G230)", key="room_input").strip()
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        total_capacity = st.number_input("Enter Total Room Capacity (for '1 to N' format)", min_value=1, max_value=2000, value=2000, key="total_capacity_input")
                    with col2:
                        capacity_per_format = st.number_input("Capacity per Format (for 'A/B' formats)", min_value=1, max_value=100, value=30, key="capacity_per_format_input")
                    
                    seat_format = st.radio("Select Seat Assignment Format:", ["1 to N", "1A to NA", "1B to NB"], key="seat_format_radio")

                    if room:
                        # Display current room status
                        room_assigned_seats_current = assigned_seats_df[
                            (assigned_seats_df["room number"] == room) & # Use lowercase
                            (assigned_seats_df["date"] == date) & # Use lowercase
                            (assigned_seats_df["shift"] == shift) # Use lowercase
                        ]["seat number"].tolist() # Use lowercase

                        a_seats_used_current = len([s for s in room_assigned_seats_current if str(s).endswith("A") and s])
                        b_seats_used_current = len([s for s in room_assigned_seats_current if str(s).endswith("B") and s])
                        no_suffix_seats_used_current = len([s for s in room_assigned_seats_current if not str(s).endswith("A") and not str(s).endswith("B")])
                        
                        st.subheader("📊 Current Room Status")
                        if seat_format in ["1A to NA", "1B to NB"]:
                            a_remaining_current = capacity_per_format - a_seats_used_current
                            b_remaining_current = capacity_per_format - b_seats_used_current
                            st.info(f"A-format: **{a_remaining_current}** remaining ({a_seats_used_current}/{capacity_per_format} used)")
                            st.info(f"B-format: **{b_remaining_current}** remaining ({b_seats_used_current}/{capacity_per_format} used)")
                            st.session_state.current_room_status_a_rem = a_remaining_current
                            st.session_state.current_room_status_b_rem = b_remaining_current
                            st.session_state.current_room_status_total_rem = None # Reset total
                        else: # 1 to N format
                            remaining_current = total_capacity - no_suffix_seats_used_current
                            st.info(f"Total: **{remaining_current}** seats remaining ({no_suffix_seats_used_current}/{total_capacity} used)")
                            st.session_state.current_room_status_total_rem = remaining_current
                            st.session_state.current_room_status_a_rem = None # Reset A/B
                            st.session_state.current_room_status_b_rem = None # Reset A/B
                    st.markdown("---")

                    if st.button("✅ Assign Seats", key="assign_button"):
                        if not room:
                            st.error("Please enter a Room Number before assigning seats.")
                            st.stop() # Stop if no room number

                        # Get all students for the selected paper from sitting plan
                        roll_cols = [col for col in sitting_plan.columns if col.lower().startswith("roll number")]
                        paper_rows = sitting_plan[sitting_plan["Paper Code"].astype(str) == paper_code]
                        all_rolls = paper_rows[roll_cols].values.flatten()
                        all_rolls = [str(r).strip() for r in all_rolls if str(r).strip() and str(r).lower() != 'nan']

                        # Get students already assigned for this paper, date, and shift
                        already_assigned_rolls = assigned_seats_df[
                            (assigned_seats_df["paper code"].astype(str) == paper_code) & # Use lowercase
                            (assigned_seats_df["date"] == date) & # Use lowercase
                            (assigned_seats_df["shift"] == shift) # Use lowercase
                        ]["roll number"].astype(str).tolist() # Use lowercase

                        unassigned_rolls = [r for r in all_rolls if r not in already_assigned_rolls]

                        if not unassigned_rolls:
                            st.warning("⚠️ All students for this paper are already assigned for this date/shift!")
                            st.stop()

                        suffix = ""
                        format_capacity_for_assignment = 0
                        if seat_format == "1 to N":
                            suffix = ""
                            format_capacity_for_assignment = total_capacity
                        elif seat_format == "1A to NA":
                            suffix = "A"
                            format_capacity_for_assignment = capacity_per_format
                        elif seat_format == "1B to NB":
                            suffix = "B"
                            format_capacity_for_assignment = capacity_per_format
                        
                        # Get currently occupied physical seats in this room for this session
                        # Use tuple for key: (room_num, seat_num, date, shift)
                        occupied_physical_seat_keys = set(
                            (str(x[0]), str(x[1]), str(x[2]), str(x[3])) for x in assigned_seats_df[
                                (assigned_seats_df["room number"] == room) & # Use lowercase
                                (assigned_seats_df["date"] == date) & # Use lowercase
                                (assigned_seats_df["shift"] == shift) # Use lowercase
                            ][['room number', 'seat number', 'date', 'shift']].values # Use lowercase
                        )
                        
                        available_seat_numbers = [] # List of numeric part of seats that are free
                        for i in range(1, format_capacity_for_assignment + 1):
                            prospective_seat_string = f"{i}{suffix}"
                            prospective_seat_key = (str(room), prospective_seat_string, str(date), str(shift))
                            if prospective_seat_key not in occupied_physical_seat_keys:
                                available_seat_numbers.append(i)

                        if not available_seat_numbers:
                            st.error(f"❌ ERROR: No seats available in **{seat_format}** format for Room {room}! Please manually change to a different format (e.g., '1A to NA' or '1B to NB') or room.")
                            st.stop()

                        if len(available_seat_numbers) < len(unassigned_rolls):
                            st.warning(f"⚠️ Capacity Warning: Only **{len(available_seat_numbers)}** seats available in **{seat_format}** format, but **{len(unassigned_rolls)}** students need assignment.")
                            st.warning(f"💡 This will assign the first **{len(available_seat_numbers)}** students. Remaining students will need assignment in a different format or room.")

                        seats_to_assign_count = min(len(available_seat_numbers), len(unassigned_rolls))
                        assigned_seat_strings = [f"{available_seat_numbers[i]}{suffix}" for i in range(seats_to_assign_count)]
                        students_to_assign = unassigned_rolls[:seats_to_assign_count]

                        assigned_rows = []
                        for i, roll in enumerate(students_to_assign):
                            seat_num_str = assigned_seat_strings[i]
                            current_assignment_key = (str(room), seat_num_str, str(date), str(shift))
                            
                            # Double check for real-time conflicts (should be rare with `occupied_physical_seat_keys`)
                            if current_assignment_key in occupied_physical_seat_keys:
                                st.warning(f"⚠️ Conflict: Seat **{seat_num_str}** in Room **{room}** is already assigned for this date/shift. Skipping assignment for Roll Number **{roll}**.")
                            else:
                                assigned_rows.append({
                                    "Roll Number": roll,
                                    "Paper Code": paper_code,
                                    "Paper Name": paper_name,
                                    "Room Number": room,
                                    "Seat Number": seat_num_str,
                                    "Date": date,
                                    "Shift": shift
                                })
                                occupied_physical_seat_keys.add(current_assignment_key) # Add to set to prevent re-assignment in this session

                        new_assignments_df = pd.DataFrame(assigned_rows)

                        if new_assignments_df.empty:
                            st.warning("No new unique seats could be assigned in this attempt, possibly due to conflicts with existing assignments.")
                        else:
                            # Concatenate and drop duplicates (in case of re-assignment logic issues)
                            assigned_seats_df = pd.concat([assigned_seats_df, new_assignments_df], ignore_index=True)
                            # Keep only unique Roll Number-Paper Code-Date-Shift assignments (ensures a student is assigned to only one seat for an exam)
                            assigned_seats_df.drop_duplicates(subset=["Roll Number", "Paper Code", "Date", "Shift"], inplace=True)
                            
                            # Save the updated assigned seats DataFrame locally
                            success, msg = save_uploaded_file(assigned_seats_df, ASSIGNED_SEATS_FILE)
                            if success:
                                st.success(f"✅ Successfully assigned **{len(new_assignments_df)}** students to Room **{room}** using **{seat_format}** format.")
                                st.dataframe(new_assignments_df) # Display the newly assigned students
                                
                                # Upload to Supabase
                                supabase_success, supabase_msg = upload_csv_to_supabase("assigned_seats", get_center_filepath(ASSIGNED_SEATS_FILE))
                                if supabase_success:
                                    st.success(f"Supabase: {supabase_msg}")
                                    load_data() # Reload all data to refresh state
                                    st.rerun() # Rerun to refresh the UI and current_room_status
                                else:
                                    st.error(f"Supabase Upload Error: {supabase_msg}")
                            else:
                                st.error(f"Error saving assigned seats: {msg}")

                            # Update displayed room status after assignment
                            st.subheader("📊 Updated Room Status")
                            updated_room_assigned_seats = assigned_seats_df[
                                (assigned_seats_df["room number"] == room) & # Use lowercase
                                (assigned_seats_df["date"] == date) & # Use lowercase
                                (assigned_seats_df["shift"] == shift) # Use lowercase
                            ]["seat number"].tolist() # Use lowercase

                            updated_a_seats_used = len([s for s in updated_room_assigned_seats if str(s).endswith("A")])
                            updated_b_seats_used = len([s for s in updated_room_assigned_seats if str(s).endswith("B")])
                            updated_no_suffix_seats_used = len([s for s in updated_room_assigned_seats if not str(s).endswith("A") and not str(s).endswith("B")])

                            if seat_format in ["1A to NA", "1B to NB"]:
                                updated_a_remaining = capacity_per_format - updated_a_seats_used
                                updated_b_remaining = capacity_per_format - updated_b_seats_used
                                st.info(f"A-format: **{updated_a_remaining}** remaining ({updated_a_seats_used}/{capacity_per_format} used)")
                                st.info(f"B-format: **{updated_b_remaining}** remaining ({updated_b_seats_used}/{capacity_per_format} used)")
                                st.session_state.current_room_status_a_rem = updated_a_remaining
                                st.session_state.current_room_status_b_rem = updated_b_remaining
                            else: # 1 to N format
                                updated_total_remaining = total_capacity - updated_no_suffix_seats_used
                                st.info(f"Total: **{updated_total_remaining}** seats remaining ({updated_no_suffix_seats_used}/{total_capacity} used)")
                                st.session_state.current_room_status_total_rem = updated_total_remaining
        
        elif admin_option == "Room Occupancy Report":
            display_room_occupancy_report(sitting_plan, assigned_seats_df, timetable)

        elif admin_option == "Room Chart Report":
            st.subheader("Generate Room Chart")
            if sitting_plan.empty or assigned_seats_df.empty or timetable.empty:
                st.warning("Please upload Sitting Plan, Assigned Seats, and Timetable files to generate room chart.")
            else:
                selected_chart_date = st.selectbox("Select Date", sorted(timetable['Date'].dropna().unique().tolist()), key="cs_room_chart_date")
                selected_chart_shift = st.selectbox("Select Shift", sorted(timetable['Shift'].dropna().unique().tolist()), key="cs_room_chart_shift")

                if st.button("Generate Room Chart"):
                    with st.spinner("Generating room chart..."):
                        # The generate_room_chart_report function now returns a string message if there's an error
                        room_chart_output = generate_room_chart_report(selected_chart_date, selected_chart_shift, sitting_plan, assigned_seats_df, timetable)
                        
                        # Check if the output is an error message (string) or the actual chart data
                        if room_chart_output and "Error:" in room_chart_output:
                            st.error(room_chart_output) # Display the error message
                        elif room_chart_output:
                            st.text_area("Generated Room Chart", room_chart_output, height=600)
                            
                            # Download button
                            file_name = f"room_chart_{selected_chart_date}_{selected_chart_shift}.csv"
                            st.download_button(
                                label="Download Room Chart as CSV",
                                data=room_chart_output.encode('utf-8'),
                                file_name=file_name,
                                mime="text/csv",
                            )
                        else:
                            st.warning("Could not generate room chart. Please ensure data is complete and assignments are made.")

        elif admin_option == "Generate College Statistics":
            st.subheader("📈 Generate College Statistics")
            attestation_filepath = get_center_filepath(ATTESTATION_DATA_FILE)
            output_stats_filepath = get_center_filepath(COLLEGE_STATISTICS_FILE)

            if not os.path.exists(attestation_filepath) or os.stat(attestation_filepath).st_size == 0:
                st.warning(f"Attestation data file not found or is empty at `{attestation_filepath}`. Please process attestation PDFs or download data first.")
            else:
                st.info(f"Using attestation data from: `{attestation_filepath}`")
                if st.button("Generate College Statistics Report"):
                    with st.spinner("Generating statistics..."):
                        success, message = generate_college_statistics(attestation_filepath, output_stats_filepath)
                        if success:
                            st.success(message)
                            if os.path.exists(output_stats_filepath):
                                with open(output_stats_filepath, "rb") as f:
                                    st.download_button(
                                        label="Download College Statistics CSV",
                                        data=f.read(),
                                        file_name=COLLEGE_STATISTICS_FILE,
                                        mime="text/csv"
                                    )
                        else:
                            st.error(message)

        elif admin_option == "Remuneration Bill Generation":
            st.subheader("💰 Remuneration Bill Generation")
            
            # Load required dataframes
            sitting_plan, timetable, assigned_seats_df, attestation_df = load_data()
            shift_assignments_df = load_shift_assignments()
            room_invigilator_assignments_df = load_room_invigilator_assignments()
            exam_team_members = load_exam_team_members()

            if shift_assignments_df.empty:
                st.warning("Shift assignments data is required to calculate remuneration. Please upload 'shift_assignments.csv'.")
                st.stop()
            if room_invigilator_assignments_df.empty:
                st.warning("Room invigilator assignments data is required for comprehensive remuneration. Please upload 'room_invigilator_assignments.csv'.")
            if timetable.empty:
                st.warning("Timetable data is required for remuneration calculations. Please upload 'timetable.csv'.")
            if assigned_seats_df.empty:
                st.warning("Assigned seats data is required for class worker remuneration. Please assign seats or upload 'assigned_seats.csv'.")


            st.markdown("---")
            st.subheader("Manual Remuneration Rates")
            # Default rates (can be moved to st.secrets for production)
            default_rates = {
                'senior_center_superintendent_rate': 1000,
                'center_superintendent_rate': 800,
                'assistant_center_superintendent_rate': 600,
                'permanent_invigilator_rate': 500,
                'assistant_permanent_invigilator_rate': 400,
                'invigilator_rate': 300,
                'conveyance_rate': 100,
                'class_3_worker_rate_per_student': 5, # Example rate
                'class_4_worker_rate_per_student': 3, # Example rate
                'holiday_conveyance_allowance_rate': 200 # New rate for holiday duty conveyance
            }
            manual_rates = {}
            for key, default_val in default_rates.items():
                manual_rates[key] = st.number_input(f"{key.replace('_', ' ').title()}", value=float(default_val), key=f"rate_{key}")

            st.markdown("---")
            st.subheader("Preparation & Closing Day Assignments")
            st.info("Assign roles and dates for preparation and closing duties. This is used for additional remuneration.")
            
            prep_closing_assignments = {}
            # Allow adding multiple prep/closing assignments
            num_prep_closing_entries = st.number_input("Number of Prep/Closing Assignments", min_value=0, value=1, key="num_prep_closing_entries")
            for i in range(num_prep_closing_entries):
                st.markdown(f"**Assignment {i+1}**")
                member_name = st.selectbox(f"Select Team Member {i+1}", [''] + sorted(exam_team_members), key=f"prep_close_member_{i}")
                
                if member_name:
                    role_options = list(remuneration_rules.keys())
                    assigned_role = st.selectbox(f"Assigned Role for {member_name}", [''] + role_options, key=f"prep_close_role_{i}")
                    
                    prep_days_input = st.text_area(f"Preparation Days (DD-MM-YYYY, comma-separated for {member_name})", key=f"prep_days_{i}").strip()
                    closing_days_input = st.text_area(f"Closing Days (DD-MM-YYYY, comma-separated for {member_name})", key=f"closing_days_{i}").strip()
                    
                    prep_days = [d.strip() for d in prep_days_input.split(',') if d.strip()] if prep_days_input else []
                    closing_days = [d.strip() for d in closing_days_input.split(',') if d.strip()] if closing_days_input else []

                    if member_name:
                        prep_closing_assignments[member_name] = {
                            'role': assigned_role,
                            'prep_days': prep_days,
                            'closing_days': closing_days
                        }
            st.markdown("---")
            
            st.subheader("Holiday Dates")
            holiday_dates_input = st.text_area("Enter Holiday Dates (DD-MM-YYYY, comma-separated)", key="holiday_dates_input").strip()
            holiday_dates = [d.strip() for d in holiday_dates_input.split(',') if d.strip()] if holiday_dates_input else []
            
            st.markdown("---")

            st.subheader("Filter by Exam Classes for Bill Generation")
            unique_classes_in_timetable = sorted(timetable['Class'].astype(str).dropna().unique().tolist())
            selected_classes_for_bill = st.multiselect(
                "Select specific exam classes to calculate bills for (leave empty for all classes)",
                unique_classes_in_timetable,
                key="selected_classes_for_bill"
            )

            if st.button("Generate Remuneration Bills"):
                with st.spinner("Calculating remuneration..."):
                    df_individual_bills, df_role_summary_matrix, df_class_3_4_final_bills = calculate_remuneration(
                        shift_assignments_df,
                        room_invigilator_assignments_df,
                        timetable,
                        assigned_seats_df,
                        manual_rates,
                        prep_closing_assignments,
                        holiday_dates,
                        selected_classes_for_bill # Pass selected classes
                    )

                    st.success("Remuneration bills generated!")

                    st.subheader("Individual Bills")
                    if not df_individual_bills.empty:
                        df_individual_bills_total = add_total_row(df_individual_bills)
                        st.dataframe(df_individual_bills_total)
                    else:
                        st.info("No individual bills to display.")

                    st.subheader("Role Summary Matrix by Date")
                    if not df_role_summary_matrix.empty:
                        df_role_summary_matrix_total = df_role_summary_matrix # Total row is already in the matrix
                        st.dataframe(df_role_summary_matrix_total)
                    else:
                        st.info("No role summary to display.")

                    st.subheader("Class 3 & Class 4 Worker Bills")
                    if not df_class_3_4_final_bills.empty:
                        df_class_3_4_final_bills_total = add_total_row(df_class_3_4_final_bills)
                        st.dataframe(df_class_3_4_final_bills_total)
                    else:
                        st.info("No Class 3 & Class 4 worker bills to display.")

                    if not df_individual_bills.empty or not df_role_summary_matrix.empty or not df_class_3_4_final_bills.empty:
                        excel_file, excel_filename = save_bills_to_excel(df_individual_bills, df_role_summary_matrix, df_class_3_4_final_bills, f"remuneration_bills_{st.session_state.center_id}.xlsx")
                        st.download_button(
                            label="Download All Bills as Excel",
                            data=excel_file,
                            file_name=excel_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.warning("No data generated for any bill type to download.")
        
        elif admin_option == "Report Panel":
            display_report_panel()
        
    else:
        st.warning("Enter valid Admin credentials.")

elif menu == "Centre Superintendent Panel":
    st.subheader("🔐 Centre Superintendent Login")
    if cs_login():
        st.success("Login successful!")
        
        # Load relevant data for CS panel
        sitting_plan, timetable, assigned_seats_df, attestation_df = load_data()
        all_reports_df = load_cs_reports_csv()

        st.markdown("---")
        st.subheader("Submit Daily Report")
        
        # Get unique dates and shifts from timetable for selections
        report_date_options = sorted(timetable['Date'].dropna().unique().tolist())
        report_shift_options = sorted(timetable['Shift'].dropna().unique().tolist())

        if not report_date_options or not report_shift_options:
            st.warning("Timetable data is empty or incomplete. Please ask the Admin to upload it.")
            st.stop()
        
        selected_report_date = st.selectbox("Select Date", report_date_options, key="cs_report_date")
        selected_report_shift = st.selectbox("Select Shift", report_shift_options, key="cs_report_shift")

        # Filter rooms based on assigned students for the selected date/shift
        filtered_assigned_seats_for_session = assigned_seats_df[
            (assigned_seats_df['date'] == selected_report_date) & 
            (assigned_seats_df['shift'] == selected_report_shift)
        ]
        room_options = sorted(filtered_assigned_seats_for_session['room number'].dropna().unique().tolist())

        if not room_options:
            st.warning("No students assigned to rooms for the selected date and shift. Please assign seats first.")
            st.stop()

        selected_room_num = st.selectbox("Select Room Number", room_options, key="cs_report_room")
        
        # Get paper codes and names for the selected room, date, and shift
        papers_in_room_session = filtered_assigned_seats_for_session[
            filtered_assigned_seats_for_session['room number'] == selected_room_num
        ][['paper code', 'paper name']].drop_duplicates()
        
        paper_display_options = [f"{row['paper code']} - {row['paper name']}" for index, row in papers_in_room_session.iterrows()]
        
        if not paper_display_options:
            st.warning("No papers assigned to students in this room for the selected date and shift.")
            st.stop()

        selected_paper_display = st.selectbox("Select Paper Code and Name", paper_display_options, key="cs_report_paper")
        selected_paper_code = selected_paper_display.split(' - ')[0].strip()
        selected_paper_name = selected_paper_display.split(' - ', 1)[1].strip()

        # Get the class for the selected paper/date/shift from the timetable
        matching_timetable_entry = timetable[(timetable['Date'].astype(str) == selected_report_date) & 
                                             (timetable['Shift'].astype(str) == selected_report_shift) &
                                             (timetable['Paper Code'].astype(str) == selected_paper_code) &
                                             (timetable['Paper Name'].astype(str) == selected_paper_name)]
        selected_class = matching_timetable_entry['Class'].iloc[0] if not matching_timetable_entry.empty else "N/A"
        st.info(f"Detected Class for this paper: **{selected_class}**")


        # Generate a unique report key
        report_key = f"{selected_report_date}_{selected_report_shift}_{selected_room_num}_{selected_paper_code}"
        
        # Load existing report data if available for pre-filling
        report_data_exists, existing_report = load_single_cs_report_csv(report_key)

        st.markdown("---")
        st.subheader("Report Details")

        # Absent roll numbers
        current_absent_rolls = existing_report.get('absent_roll_numbers', [])
        absent_rolls_input = st.text_area(
            "Enter Absent Roll Numbers (comma-separated)",
            value=", ".join(current_absent_rolls),
            key="absent_rolls_input"
        )
        absent_roll_numbers = [r.strip() for r in absent_rolls_input.split(',') if r.strip()]

        # UFM roll numbers
        current_ufm_rolls = existing_report.get('ufm_roll_numbers', [])
        ufm_rolls_input = st.text_area(
            "Enter UFM Roll Numbers (comma-separated)",
            value=", ".join(current_ufm_rolls),
            key="ufm_rolls_input"
        )
        ufm_roll_numbers = [r.strip() for r in ufm_rolls_input.split(',') if r.strip()]

        if st.button("Submit Daily Report"):
            report_data = {
                'report_key': report_key,
                'date': selected_report_date,
                'shift': selected_report_shift,
                'room_num': selected_room_num,
                'paper_code': selected_paper_code,
                'paper_name': selected_paper_name,
                'class': selected_class, # Save the detected class
                'absent_roll_numbers': absent_roll_numbers,
                'ufm_roll_numbers': ufm_roll_numbers
            }
            success, msg = save_cs_report_csv(report_key, report_data)
            if success:
                st.success(msg)
                # Upload to Supabase after saving locally
                supabase_success, supabase_msg = upload_csv_to_supabase("cs_reports", get_center_filepath(CS_REPORTS_FILE))
                if supabase_success:
                    st.success(f"Supabase: {supabase_msg}")
                    # No rerun here to allow user to generate UFM form
                else:
                    st.error(f"Supabase Upload Error: {supabase_msg}")
            else:
                st.error(msg)
        
        st.markdown("---")
        st.subheader("Generate UFM Print Form")
        if not ufm_roll_numbers:
            st.info("No UFM roll numbers entered in the report above to generate a print form.")
        else:
            selected_ufm_roll_number = st.selectbox("Select UFM Roll Number to Print Form", ufm_roll_numbers, key="ufm_form_roll")
            if st.button(f"Generate UFM Form for {selected_ufm_roll_number}"):
                if attestation_df.empty:
                    st.warning("Attestation data is required to generate UFM forms. Please ask Admin to upload/download it.")
                else:
                    ufm_form_output = generate_ufm_print_form(
                        selected_ufm_roll_number, 
                        attestation_df, 
                        assigned_seats_df, 
                        timetable, 
                        selected_report_date, 
                        selected_report_shift, 
                        selected_paper_code, 
                        selected_paper_name
                    )
                    st.text_area("UFM Print Form", ufm_form_output, height=600)
                    file_name = f"UFM_Form_{selected_ufm_roll_number}_{selected_report_date}.txt"
                    st.download_button(
                        label="Download UFM Form as TXT",
                        data=ufm_form_output.encode('utf-8'),
                        file_name=file_name,
                        mime="text/plain",
                    )
    else:
        st.warning("Enter valid Centre Superintendent credentials.")
